# -*- coding: utf-8 -*-
#!/usr/bin/env python3
"""
Data Collection Portal  v3
Admin: admin / admin123
"""
import os, io, json, re, uuid
from datetime import datetime
from functools import wraps
from flask import (Flask, render_template_string, request, redirect,
                   url_for, session, flash, send_file, jsonify)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import sqlite3, openpyxl, zipfile
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(BASE, 'data')
PDFS = os.path.join(DATA, 'pdfs')
DB   = os.path.join(DATA, 'portal.db')
TEMP = os.path.join(DATA, 'tmp')
for d in [DATA, PDFS, TEMP]: os.makedirs(d, exist_ok=True)

def save_pending(data):
    fname = 'pending_' + uuid.uuid4().hex + '.json'
    with open(os.path.join(TEMP, fname), 'w') as ff:
        json.dump(data, ff)
    return fname

def load_pending():
    fname = session.get('pending_file')
    if not fname: return None
    fpath = os.path.join(TEMP, fname)
    if not os.path.exists(fpath): return None
    with open(fpath, 'r') as ff:
        return json.load(ff)

def clear_pending():
    fname = session.pop('pending_file', None)
    if fname:
        try: os.remove(os.path.join(TEMP, fname))
        except: pass

app = Flask(__name__)
app.secret_key = 'cgtmse-v3-2026-xK9mP'
app.config['MAX_CONTENT_LENGTH'] = 25 * 1024 * 1024

# ══════════════════════════════════════════════════════════════════════════════
# DATABASE
# ══════════════════════════════════════════════════════════════════════════════
SCHEMA = """
CREATE TABLE IF NOT EXISTS admins (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    username      TEXT UNIQUE NOT NULL,
    password_hash TEXT NOT NULL,
    full_name     TEXT DEFAULT 'Admin',
    created_at    TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS branches (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    branch_code   TEXT UNIQUE NOT NULL,
    branch_name   TEXT,
    ro            TEXT,
    zo            TEXT,
    email         TEXT,
    password_hash TEXT,
    first_login   INTEGER DEFAULT 1
);
CREATE TABLE IF NOT EXISTS uploads (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    filename      TEXT NOT NULL,
    week_label    TEXT NOT NULL,
    row_count     INTEGER DEFAULT 0,
    branch_count  INTEGER DEFAULT 0,
    uploaded_by   TEXT,
    uploaded_at   TEXT DEFAULT (datetime('now','localtime')),
    col_config    TEXT DEFAULT '[]',
    mapped        INTEGER DEFAULT 0,
    active        INTEGER DEFAULT 1
);
CREATE TABLE IF NOT EXISTS loan_accounts (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    upload_id    INTEGER NOT NULL,
    week_label   TEXT NOT NULL,
    branch_code  TEXT NOT NULL,
    branch_name  TEXT,
    zo           TEXT,
    ro           TEXT,
    row_data     TEXT DEFAULT '{}',
    branch_data  TEXT DEFAULT '{}',
    pdf_files    TEXT DEFAULT '{}',
    status       TEXT DEFAULT 'Pending',
    submitted_at TEXT,
    FOREIGN KEY(upload_id) REFERENCES uploads(id)
);
"""

def db_conn():
    c = sqlite3.connect(DB)
    c.row_factory = sqlite3.Row
    c.execute("PRAGMA journal_mode=WAL")
    return c

def init_db():
    c = db_conn()
    c.executescript(SCHEMA)
    if not c.execute("SELECT 1 FROM admins LIMIT 1").fetchone():
        c.execute("INSERT INTO admins(username,password_hash,full_name) VALUES(?,?,?)",
                  ('admin', generate_password_hash('admin123'), 'Administrator'))
        c.commit()
    c.close()

init_db()

# ══════════════════════════════════════════════════════════════════════════════
# COLUMN INTELLIGENCE
# ══════════════════════════════════════════════════════════════════════════════
ZO_ALIASES  = ['zo','zonal office','zonal','zone','z.o','zone name','zo name']
RO_ALIASES  = ['ro','regional office','regional','region','r.o','region name','ro name']
BC_ALIASES  = ['branch code','br code','br. code','sol id','sol_id','branch id','br_code',
               'sol_code','brnch code','branch_code']
BN_ALIASES  = ['branch name','br name','branch','name of branch']

# Keywords that should always be LOCKED — financial/reference data
ALWAYS_LOCKED_KW = [
    'amount','amt','sanction','sanctioned','approved','limit','outstanding',
    'balance','disburs','overdue','npa','rate','roi','interest','emi',
    'loan no','loan account','account no','a/c','ac no','acct',
    'date','tenure','period','installment','repay','scheme','product',
    'cgtmse','guarantee','coverage','fee','premium','charge',
    'pan','aadhar','gstin','cin','ifsc','micr',
]

def is_always_locked(header):
    h = str(header).strip().lower()
    return any(kw in h for kw in ALWAYS_LOCKED_KW)

def h_match(h, aliases):
    import re as _re
    h = str(h).strip().lower()
    # Exact match first
    if h in aliases:
        return True
    # Word-boundary match: alias must appear as a whole word, not substring of another word
    # e.g. 'ro' should NOT match 'approved', 'borrower', 'product'
    for a in aliases:
        if len(a) <= 3:
            # Short aliases (zo, ro, etc.) — require word boundary
            pattern = r'(?<![a-z])' + _re.escape(a) + r'(?![a-z])'
            if _re.search(pattern, h):
                return True
        else:
            # Longer aliases — substring match is fine
            if a in h:
                return True
    return False

def classify_known(header):
    if h_match(header, ZO_ALIASES):  return 'zo'
    if h_match(header, RO_ALIASES):  return 'ro'
    if h_match(header, BC_ALIASES):  return 'branch_code'
    if h_match(header, BN_ALIASES):  return 'branch_name'
    return None

def detect_structure(headers, data_rows):
    """Return col_config: list of dicts with full metadata per column."""
    config = []
    for i, h in enumerate(headers):
        if h is None or str(h).strip() == '': continue
        h = str(h).strip()
        known   = classify_known(h)
        filled  = sum(1 for r in data_rows
                      if i < len(r) and r[i] is not None and str(r[i]).strip() != '')
        ratio   = round(filled / len(data_rows), 2) if data_rows else 0
        # Default type: locked if mostly filled, fillable if mostly empty
        default = 'locked' if ratio >= 0.4 else 'fillable'
        # System columns always locked
        if known in ('zo','ro','branch_code','branch_name'): default = 'locked'
        # Financial/reference columns always locked regardless of fill ratio
        if is_always_locked(h): default = 'locked'
        config.append({'idx': i, 'name': h, 'known': known,
                       'fill_ratio': ratio, 'type': default})
    return config

def safe(row, idx, default=''):
    try:
        v = row[idx] if (idx is not None and 0 <= idx < len(row)) else default
        return str(v).strip() if v is not None else default
    except: return default

def week_now():
    n = datetime.now()
    return f"Sheet {n.isocalendar()[1]} · {n.strftime('%b %Y')}"

# ══════════════════════════════════════════════════════════════════════════════
# AUTH
# ══════════════════════════════════════════════════════════════════════════════
def admin_only(f):
    @wraps(f)
    def w(*a,**k):
        if session.get('role')!='admin': return redirect(url_for('admin_login'))
        return f(*a,**k)
    return w

def branch_only(f):
    @wraps(f)
    def w(*a,**k):
        if session.get('role')!='branch': return redirect(url_for('branch_login'))
        return f(*a,**k)
    return w

# ══════════════════════════════════════════════════════════════════════════════
# INDEX
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/')
def index():
    if session.get('role')=='admin':  return redirect(url_for('dashboard'))
    if session.get('role')=='branch': return redirect(url_for('branch_home'))
    return redirect(url_for('admin_login'))

# ══════════════════════════════════════════════════════════════════════════════
# ADMIN LOGIN
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/admin/login', methods=['GET','POST'])
def admin_login():
    if request.method=='POST':
        u = request.form.get('username','').strip().lower()
        p = request.form.get('password','')
        c = db_conn()
        row = c.execute("SELECT * FROM admins WHERE username=?",(u,)).fetchone()
        c.close()
        if row and check_password_hash(row['password_hash'],p):
            session.clear()
            session.update({'role':'admin','username':u,'full_name':row['full_name']})
            return redirect(url_for('dashboard'))
        flash('Wrong username or password.','danger')
    return render_template_string(T_LOGIN)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('admin_login'))

# ══════════════════════════════════════════════════════════════════════════════
# UPLOAD — Step 1: pick file
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/upload', methods=['GET','POST'])
@admin_only
def upload():
    if request.method=='POST':
        f    = request.files.get('file')
        week = request.form.get('week_label', week_now()).strip()
        if not f or not re.search(r'\.xlsx?$', f.filename or '', re.I):
            flash('Please upload a .xlsx or .xls file.','danger')
            return redirect(url_for('upload'))
        try:
            wb = openpyxl.load_workbook(io.BytesIO(f.read()),
                                        read_only=True, data_only=True)
            ws = wb.active
            all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
            wb.close()
        except Exception as e:
            flash(f'Cannot read file: {e}','danger')
            return redirect(url_for('upload'))
        if len(all_rows) < 2:
            flash('File has no data rows.','warning')
            return redirect(url_for('upload'))

        headers   = [str(h).strip() if h else '' for h in all_rows[0]]
        data_rows = all_rows[1:]
        col_cfg   = detect_structure(headers, data_rows)

        # Validate at least ZO or RO
        has_zo = any(c['known']=='zo' for c in col_cfg)
        has_ro = any(c['known']=='ro' for c in col_cfg)
        if not has_zo and not has_ro:
            flash(f'File must have ZO or RO column. Found: {", ".join(h for h in headers if h)}','danger')
            return redirect(url_for('upload'))

        # Store in temp file — avoids 4KB cookie size limit on large files
        pfname = save_pending({'filename':f.filename,'week':week,'headers':headers,
            'data':[[str(v) if v is not None else '' for v in r] for r in data_rows],
            'col_cfg':col_cfg})
        session['pending_file'] = pfname
        return redirect(url_for('map_fields'))

    c = db_conn()
    recent = c.execute("""SELECT u.*,
        (SELECT COUNT(*) FROM loan_accounts WHERE upload_id=u.id) AS accounts
        FROM uploads u ORDER BY uploaded_at DESC LIMIT 20""").fetchall()
    c.close()
    return render_template_string(T_UPLOAD, recent=recent, week_now=week_now())

# ══════════════════════════════════════════════════════════════════════════════
# MAP FIELDS — Step 2: admin reviews/edits column types
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/map', methods=['GET','POST'])
@admin_only
def map_fields():
    pending = load_pending()
    if not pending:
        flash('Upload session expired. Please upload again.','warning')
        return redirect(url_for('upload'))

    if request.method=='POST':
        col_cfg = pending['col_cfg']
        for cc in col_cfg:
            override = request.form.get(f"col_{cc['idx']}")
            if override in ('locked','fillable','pdf'):
                cc['type'] = override
            # Extra options only apply to fillable columns
            if cc['type'] == 'fillable':
                cc['format']    = request.form.get(f"fmt_{cc['idx']}", 'text')
                cc['required']  = request.form.get(f"req_{cc['idx']}") == '1'
                raw_opts        = request.form.get(f"opts_{cc['idx']}", '').strip()
                cc['options']   = [o.strip() for o in raw_opts.split(',') if o.strip()] if raw_opts else []
            else:
                cc['format']   = 'text'
                cc['required'] = False
                cc['options']  = []
        pending['col_cfg'] = col_cfg
        fname = session.get('pending_file')
        if fname:
            with open(os.path.join(TEMP, fname), 'w') as ff:
                json.dump(pending, ff)
        return redirect(url_for('save_upload'))

    return render_template_string(T_MAP_FIELDS, pending=pending)

# ══════════════════════════════════════════════════════════════════════════════
# SAVE UPLOAD — Step 3: process and save to DB
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/save_upload')
@admin_only
def save_upload():
    pending = load_pending()
    clear_pending()
    if not pending:
        flash('Upload session expired. Please upload again.','warning')
        return redirect(url_for('upload'))

    headers   = pending['headers']
    data_rows = pending['data']
    col_cfg   = pending['col_cfg']
    week      = pending['week']
    filename  = pending['filename']

    db = db_conn()

    # RO→ZO master map
    ro_zo = {r['ro'].strip().lower(): r['zo']
             for r in db.execute(
                 "SELECT ro,zo FROM branches WHERE zo IS NOT NULL AND zo!=''").fetchall()
             if r['ro']}

    # Insert upload record
    cur = db.execute(
        "INSERT INTO uploads(filename,week_label,uploaded_by,col_config,mapped) VALUES(?,?,?,?,1)",
        (filename, week, session['username'], json.dumps(col_cfg)))
    uid = cur.lastrowid

    # Index of system columns
    zo_idx = next((c['idx'] for c in col_cfg if c['known']=='zo'),  -1)
    ro_idx = next((c['idx'] for c in col_cfg if c['known']=='ro'),  -1)
    bc_idx = next((c['idx'] for c in col_cfg if c['known']=='branch_code'), -1)
    bn_idx = next((c['idx'] for c in col_cfg if c['known']=='branch_name'), -1)

    inserted, branch_codes = 0, set()
    for row in data_rows:
        if not any(v for v in row if v): continue

        zo = safe(row, zo_idx)
        ro = safe(row, ro_idx)
        bc = safe(row, bc_idx)
        bn = safe(row, bn_idx)
        if not ro and not bc: continue

        if not zo and ro:
            zo = ro_zo.get(ro.strip().lower(), '')

        # Build row_data dict: ALL columns stored as-is
        row_data = {}
        for cc in col_cfg:
            row_data[cc['name']] = safe(row, cc['idx'])

        # Auto-register branch
        if bc and bc not in branch_codes:
            branch_codes.add(bc)
            ex = db.execute("SELECT id,zo FROM branches WHERE branch_code=?",(bc,)).fetchone()
            if not ex:
                db.execute("INSERT INTO branches(branch_code,branch_name,ro,zo) VALUES(?,?,?,?)",
                           (bc, bn or bc, ro, zo))
            elif zo and not ex['zo']:
                db.execute("UPDATE branches SET zo=? WHERE branch_code=?",(zo,bc))

        db.execute("""INSERT INTO loan_accounts
            (upload_id,week_label,branch_code,branch_name,zo,ro,row_data)
            VALUES(?,?,?,?,?,?,?)""",
            (uid, week, bc or ro or 'UNKNOWN', bn, zo, ro, json.dumps(row_data)))
        inserted += 1

    db.execute("UPDATE uploads SET row_count=?,branch_count=? WHERE id=?",
               (inserted, len(branch_codes), uid))
    db.commit()

    # Optional: detect missing ZO and offer to fix
    missing_ro_zo = db.execute(
        "SELECT DISTINCT ro FROM loan_accounts "
        "WHERE upload_id=? AND (zo IS NULL OR zo='') AND ro IS NOT NULL AND ro!=''",
        (uid,)).fetchall()
    missing_bc_zo = db.execute(
        "SELECT DISTINCT branch_code,branch_name,ro FROM loan_accounts "
        "WHERE upload_id=? AND (zo IS NULL OR zo='') AND branch_code IS NOT NULL",
        (uid,)).fetchall()
    db.close()

    if missing_ro_zo:
        session['fix_uid']  = uid
        session['fix_week'] = week
        session['fix_fname']= filename
        flash(f'OK {filename} — {inserted} accounts, {len(branch_codes)} branches loaded. '
              f'️ Zone (ZO) is missing for {len(missing_ro_zo)} region(s). '
              f'Assign below or skip.', 'warning')
        return redirect(url_for('fix_zones'))

    flash(f'OK {filename} — {inserted} accounts across {len(branch_codes)} branches.', 'success')
    return redirect(url_for('dashboard', week=week))

# ══════════════════════════════════════════════════════════════════════════════
# DELETE UPLOAD
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/upload/delete/<int:uid>', methods=['POST'])
@admin_only
def delete_upload(uid):
    db = db_conn()
    db.execute("DELETE FROM loan_accounts WHERE upload_id=?",(uid,))
    db.execute("DELETE FROM uploads WHERE id=?",(uid,))
    db.commit()
    db.close()
    flash('Upload deleted.','success')
    return redirect(url_for('upload'))

# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# FIX ZONES — optional step after upload when ZO is missing
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/fix_zones', methods=['GET','POST'])
@admin_only
def fix_zones():
    uid   = session.get('fix_uid')
    week  = session.get('fix_week','')
    fname = session.get('fix_fname','')
    if not uid:
        return redirect(url_for('dashboard'))

    db = db_conn()

    if request.method == 'POST':
        if request.form.get('skip'):
            session.pop('fix_uid', None); session.pop('fix_week', None); session.pop('fix_fname', None)
            db.close()
            flash('Zone assignment skipped. You can update later in Settings.', 'info')
            return redirect(url_for('dashboard', week=week))

        applied = 0
        for key, val in request.form.items():
            if key.startswith('zo_') and val.strip():
                ro  = key[3:]
                zo  = val.strip()
                db.execute("UPDATE loan_accounts SET zo=? WHERE upload_id=? AND ro=? AND (zo IS NULL OR zo='')", (zo, uid, ro))
                db.execute("UPDATE branches SET zo=? WHERE ro=? AND (zo IS NULL OR zo='')", (zo, ro))
                applied += 1
        db.commit()
        session.pop('fix_uid', None); session.pop('fix_week', None); session.pop('fix_fname', None)
        db.close()
        flash(f'OK Zone assigned for {applied} region(s).', 'success')
        return redirect(url_for('dashboard', week=week))

    missing = db.execute("""
        SELECT ro, COUNT(*) as accounts, COUNT(DISTINCT branch_code) as branches
        FROM loan_accounts WHERE upload_id=? AND (zo IS NULL OR zo='')
          AND ro IS NOT NULL AND ro!=''
        GROUP BY ro ORDER BY ro
    """, (uid,)).fetchall()

    all_zo = [r[0] for r in db.execute(
        "SELECT DISTINCT zo FROM loan_accounts WHERE zo IS NOT NULL AND zo!='' ORDER BY zo"
    ).fetchall()]
    zo_master = [r[0] for r in db.execute(
        "SELECT DISTINCT zo FROM branches WHERE zo IS NOT NULL AND zo!='' ORDER BY zo"
    ).fetchall()]
    all_zo = sorted(set(all_zo + zo_master))
    db.close()
    return render_template_string(T_FIX_ZONES,
        missing=missing, uid=uid, week=week, fname=fname,
        all_zo=all_zo, week_now=week_now())

# DASHBOARD — ZO ▶ RO ▶ Branch drill-down
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/dashboard')
@admin_only
def dashboard():
    week = request.args.get('week','')
    db   = db_conn()
    weeks = [r[0] for r in db.execute(
        "SELECT DISTINCT week_label FROM uploads WHERE active=1 ORDER BY uploaded_at DESC"
    ).fetchall()]
    if not week and weeks: week = weeks[0]

    tree = {}  # tree[zo][ro][bc] = {name,total,submitted,pct,status}
    kpi  = dict(zo=0,ro=0,branches=0,accounts=0,submitted=0,pending=0,sub_pct=0)

    if week:
        rows = db.execute("""
            SELECT branch_code,branch_name,zo,ro,status
            FROM loan_accounts WHERE week_label=? ORDER BY zo,ro,branch_code
        """,(week,)).fetchall()
        for r in rows:
            zo = r['zo'] or '—'; ro = r['ro'] or '—'; bc = r['branch_code']
            tree.setdefault(zo,{}).setdefault(ro,{}).setdefault(bc,
                {'name':r['branch_name'] or bc,'total':0,'submitted':0})
            tree[zo][ro][bc]['total']     += 1
            kpi['accounts'] += 1
            if r['status']=='Submitted':
                tree[zo][ro][bc]['submitted'] += 1
                kpi['submitted'] += 1

        for zo,ros in tree.items():
            for ro,brs in ros.items():
                for bc,b in brs.items():
                    p = round(b['submitted']/b['total']*100) if b['total'] else 0
                    b['pct']    = p
                    b['status'] = ('Submitted' if b['submitted']==b['total']
                                   else 'Partial' if b['submitted']>0 else 'Pending')
                    kpi['branches'] += 1
                kpi['ro'] += 1
            kpi['zo'] += 1
        kpi['pending']  = kpi['accounts'] - kpi['submitted']
        kpi['sub_pct']  = round(kpi['submitted']/kpi['accounts']*100) if kpi['accounts'] else 0

    db.close()
    return render_template_string(T_DASHBOARD,
        week=week, weeks=weeks, tree=tree, kpi=kpi, week_now=week_now())

# ── Account detail JSON (for admin modal) ──────────────────────────────────────
@app.route('/api/account/<int:aid>')
@admin_only
def api_account(aid):
    db  = db_conn()
    acc = db.execute("SELECT * FROM loan_accounts WHERE id=?",(aid,)).fetchone()
    db.close()
    if not acc: return jsonify(error='Not found'),404
    cfg = []
    # Get col_config for this account's upload
    db  = db_conn()
    up  = db.execute("SELECT col_config FROM uploads WHERE id=?",(acc['upload_id'],)).fetchone()
    db.close()
    if up: cfg = json.loads(up['col_config'] or '[]')
    rd  = json.loads(acc['row_data']   or '{}')
    bd  = json.loads(acc['branch_data']or '{}')
    pf  = json.loads(acc['pdf_files']  or '{}')
    return jsonify(bc=acc['branch_code'], bn=acc['branch_name'],
                   zo=acc['zo'], ro=acc['ro'],
                   row_data=rd, branch_data=bd, pdf_files=pf,
                   col_config=cfg, status=acc['status'],
                   submitted_at=acc['submitted_at'])

# ── Branch accounts JSON (for admin modal) ─────────────────────────────────────
@app.route('/api/branch_accounts')
@admin_only
def api_branch_accounts():
    bc   = request.args.get('bc','')
    week = request.args.get('week','')
    db   = db_conn()
    rows = db.execute("""
        SELECT l.id, l.status, l.row_data, l.branch_data, l.submitted_at,
               u.col_config
        FROM loan_accounts l
        JOIN uploads u ON l.upload_id=u.id
        WHERE l.branch_code=? AND l.week_label=?
        ORDER BY l.id
    """,(bc,week)).fetchall()
    db.close()
    accounts = []
    for r in rows:
        rd  = json.loads(r['row_data']   or '{}')
        bd  = json.loads(r['branch_data']or '{}')
        cfg = json.loads(r['col_config'] or '[]')
        accounts.append({'id':r['id'],'status':r['status'],
                         'row_data':rd,'branch_data':bd,
                         'col_config':cfg,'submitted_at':r['submitted_at']})
    return jsonify(accounts=accounts)

# ══════════════════════════════════════════════════════════════════════════════
# EXPORT
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/export', methods=['GET','POST'])
@admin_only
def export_page():
    db   = db_conn()
    weeks = [r[0] for r in db.execute(
        "SELECT DISTINCT week_label FROM uploads ORDER BY uploaded_at DESC").fetchall()]
    zo_list = [r[0] for r in db.execute(
        "SELECT DISTINCT zo FROM loan_accounts WHERE zo IS NOT NULL AND zo!='' ORDER BY zo").fetchall()]
    ro_list = [r[0] for r in db.execute(
        "SELECT DISTINCT ro FROM loan_accounts WHERE ro IS NOT NULL AND ro!='' ORDER BY ro").fetchall()]
    db.close()
    return render_template_string(T_EXPORT,
        weeks=weeks, zo_list=zo_list, ro_list=ro_list, week_now=week_now())

@app.route('/export/download')
@admin_only
def export_download():
    week    = request.args.get('week','')
    zo      = request.args.get('zo','')
    ro      = request.args.get('ro','')
    status  = request.args.get('status','')
    sheet   = request.args.get('sheet','all')
    inc_pdf = request.args.get('pdfs','0') == '1'

    db = db_conn()
    q  = "SELECT l.*,u.col_config FROM loan_accounts l JOIN uploads u ON l.upload_id=u.id WHERE 1=1"
    p  = []
    if week:   q+=" AND l.week_label=?"; p.append(week)
    if zo:     q+=" AND l.zo=?";         p.append(zo)
    if ro:     q+=" AND l.ro=?";         p.append(ro)
    if status: q+=" AND l.status=?";     p.append(status)
    rows = db.execute(q+" ORDER BY l.zo,l.ro,l.branch_code",p).fetchall()
    db.close()

    H_FILL  = PatternFill("solid", fgColor="0F172A")
    H_FONT  = Font(bold=True, color="FFFFFF", size=11)
    H_ALIGN = Alignment(horizontal='center', vertical='center')
    LK_FILL = PatternFill("solid", fgColor="F8FAFC")
    FL_FILL = PatternFill("solid", fgColor="FFFDE7")
    PD_FILL = PatternFill("solid", fgColor="F3E5F5")

    def style_header(ws):
        for cell in ws[1]:
            cell.fill=H_FILL; cell.font=H_FONT; cell.alignment=H_ALIGN
        ws.row_dimensions[1].height=22; ws.freeze_panes='A2'
        for i,col in enumerate(ws.columns,1):
            ws.column_dimensions[get_column_letter(i)].width=22

    def get_pdf_display(col_name, pf_dict):
        fname = pf_dict.get(col_name,'')
        if fname and os.path.exists(os.path.join(PDFS, fname)):
            orig = fname.split('_',3)[-1] if fname.count('_')>=3 else fname
            return f'PDF: {orig}'
        return ''

    def make_sheet(wb, data, title):
        if not data: return
        cfg     = json.loads(data[0]['col_config'] or '[]')
        non_sys = [c for c in cfg if c['known'] not in ('zo','ro','branch_code','branch_name')]
        headers = ['ZO','RO','Branch Code','Branch Name'] +                   [c['name'] for c in non_sys] + ['Status','Submitted At']
        ws = wb.create_sheet(title)
        ws.append(headers)
        style_header(ws)
        # Colour-code headers by field type
        for i, c in enumerate(non_sys, 5):
            cell = ws.cell(1, i)
            if c['type']=='fillable':
                cell.fill = PatternFill("solid", fgColor="E65100"); cell.font=Font(bold=True,color="FFFFFF",size=11)
            elif c['type']=='pdf':
                cell.fill = PatternFill("solid", fgColor="4A148C"); cell.font=Font(bold=True,color="FFFFFF",size=11)
        for r in data:
            rd  = json.loads(r['row_data']   or '{}')
            bd  = json.loads(r['branch_data']or '{}')
            pf  = json.loads(r['pdf_files']  or '{}')
            row_vals = [r['zo'], r['ro'], r['branch_code'], r['branch_name']]
            for c in non_sys:
                if c['type']=='locked':
                    row_vals.append(rd.get(c['name'],''))
                elif c['type']=='fillable':
                    row_vals.append(bd.get(c['name'],'') or rd.get(c['name'],''))
                else:
                    row_vals.append(get_pdf_display(c['name'], pf))
            row_vals += [r['status'], r['submitted_at'] or '']
            ws.append(row_vals)
            dr = ws.max_row
            for i, c in enumerate(non_sys, 5):
                cell = ws.cell(dr, i)
                if c['type']=='fillable': cell.fill=FL_FILL
                elif c['type']=='pdf':    cell.fill=PD_FILL
                else:                     cell.fill=LK_FILL

    def make_summary(wb, data):
        ws = wb.create_sheet('Summary')
        ws.append(['Zone (ZO)','Region (RO)','Branch Code','Branch Name','Total','Submitted','Pending','% Done'])
        style_header(ws)
        groups = {}
        for r in data:
            k = (r['zo'] or '—', r['ro'] or '—', r['branch_code'], r['branch_name'] or '')
            groups.setdefault(k, {'t':0,'s':0})
            groups[k]['t'] += 1
            if r['status']=='Submitted': groups[k]['s'] += 1
        for (zo,ro,bc,bn),g in sorted(groups.items()):
            ws.append([zo,ro,bc,bn,g['t'],g['s'],g['t']-g['s'],
                       f"{round(g['s']/g['t']*100)}%" if g['t'] else '0%'])

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    if sheet=='all':    make_sheet(wb, rows, 'All Accounts')
    elif sheet=='sub':  make_sheet(wb, [r for r in rows if r['status']=='Submitted'], 'Submitted')
    elif sheet=='pend': make_sheet(wb, [r for r in rows if r['status']!='Submitted'], 'Pending')
    elif sheet=='sum':  make_summary(wb, rows)
    else:
        make_sheet(wb, rows, 'All Accounts')
        make_sheet(wb, [r for r in rows if r['status']=='Submitted'], 'Submitted')
        make_sheet(wb, [r for r in rows if r['status']!='Submitted'], 'Pending')
        make_summary(wb, rows)
    if not wb.sheetnames: wb.create_sheet('No Data')

    stamp    = datetime.now().strftime('%Y%m%d_%H%M')
    safe_week= (week or 'All').replace(' ','_').replace('·','').replace('/','_')

    # Collect PDF files
    pdf_entries = []
    for r in rows:
        pf = json.loads(r['pdf_files'] or '{}')
        for col_name, fname in pf.items():
            disk = os.path.join(PDFS, fname)
            if fname and os.path.exists(disk):
                zo_s  = (r['zo']  or 'Unknown_ZO').replace('/','_').replace(' ','_')
                ro_s  = (r['ro']  or 'Unknown_RO').replace('/','_').replace(' ','_')
                bc_s  = (r['branch_code'] or 'Unknown').replace('/','_')
                col_s = col_name.replace('/','_').replace(' ','_')
                orig  = fname.split('_',3)[-1] if fname.count('_')>=3 else fname
                zip_p = f"PDFs/{zo_s}/{ro_s}/{bc_s}/{col_s}/{orig}"
                pdf_entries.append((zip_p, disk))

    if inc_pdf and pdf_entries:
        excel_buf = io.BytesIO(); wb.save(excel_buf); excel_buf.seek(0)
        zip_buf   = io.BytesIO()
        readme = (f"Data Collection Portal Export\nDatasheet: {week or 'All'}\n"
                  f"Exported: {datetime.now().strftime('%d %b %Y %H:%M')}\n"
                  f"Total PDFs: {len(pdf_entries)}\n\n"
                  f"Folder structure:\nPDFs/Zone/Region/BranchCode/ColumnName/filename\n")
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f"DataSheet_{safe_week}_{sheet}_{stamp}.xlsx", excel_buf.read())
            zf.writestr("README.txt", readme)
            for zip_path, disk_path in pdf_entries:
                zf.write(disk_path, zip_path)
        zip_buf.seek(0)
        return send_file(zip_buf, as_attachment=True,
                         download_name=f"DataSheet_{safe_week}_{sheet}_{stamp}.zip",
                         mimetype='application/zip')

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"DataSheet_{safe_week}_{sheet}_{stamp}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ══════════════════════════════════════════════════════════════════════════════
# SETTINGS — master + admins + clear
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/settings/download_master')
@admin_only
def download_master():
    db  = db_conn()
    brs = db.execute(
        "SELECT branch_code, branch_name, ro, zo, email FROM branches ORDER BY zo, ro, branch_name"
    ).fetchall()
    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Branch Master'

    # Header styling
    H_FILL  = PatternFill("solid", fgColor="0F172A")
    H_FONT  = Font(bold=True, color="FFFFFF", size=11)
    H_ALIGN = Alignment(horizontal='center', vertical='center')

    headers = ['ZO (Zone)', 'RO (Region)', 'Branch Code', 'Branch Name', 'Email']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill=H_FILL; cell.font=H_FONT; cell.alignment=H_ALIGN
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    for b in brs:
        ws.append([b['zo'] or '', b['ro'] or '', b['branch_code'],
                   b['branch_name'] or '', b['email'] or ''])

    # Column widths
    for col, width in zip(['A','B','C','D','E'], [25,25,18,30,35]):
        ws.column_dimensions[col].width = width

    # Add alternating row colour for readability
    light = PatternFill("solid", fgColor="F8FAFC")
    for i, row in enumerate(ws.iter_rows(min_row=2), 2):
        if i % 2 == 0:
            for cell in row:
                cell.fill = light

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Branch_Master_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/settings', methods=['GET','POST'])
@admin_only
def settings():
    db = db_conn()
    if request.method=='POST':
        action = request.form.get('action')

        if action=='master':
            mf = request.files.get('master_file')
            if mf and re.search(r'\.xlsx?$', mf.filename or '', re.I):
                try:
                    wb = openpyxl.load_workbook(io.BytesIO(mf.read()),
                                                read_only=True, data_only=True)
                    ws = wb.active
                    rows = [list(r) for r in ws.iter_rows(values_only=True)]
                    wb.close()
                    headers = rows[0]
                    cfg     = detect_structure(headers, rows[1:])
                    bc_idx  = next((c['idx'] for c in cfg if c['known']=='branch_code'),-1)
                    ro_idx  = next((c['idx'] for c in cfg if c['known']=='ro'),-1)
                    zo_idx  = next((c['idx'] for c in cfg if c['known']=='zo'),-1)
                    bn_idx  = next((c['idx'] for c in cfg if c['known']=='branch_name'),-1)
                    em_idx  = next((c['idx'] for c in cfg
                                   if 'email' in c['name'].lower() or 'mail' in c['name'].lower()),-1)
                    added=updated=0
                    for row in rows[1:]:
                        bc = safe(row,bc_idx); ro = safe(row,ro_idx)
                        zo = safe(row,zo_idx); bn = safe(row,bn_idx)
                        em = safe(row,em_idx) if em_idx>=0 else ''
                        if not bc and not ro: continue
                        ex = db.execute("SELECT id FROM branches WHERE branch_code=?",(bc,)).fetchone()
                        if ex:
                            db.execute("UPDATE branches SET branch_name=?,ro=?,zo=?,email=? WHERE branch_code=?",
                                       (bn,ro,zo,em,bc)); updated+=1
                        else:
                            db.execute("INSERT INTO branches(branch_code,branch_name,ro,zo,email) VALUES(?,?,?,?,?)",
                                       (bc,bn,ro,zo,em)); added+=1
                    db.commit()
                    flash(f'OK Master uploaded: {added} new, {updated} updated.','success')
                except Exception as e:
                    flash(f'Error: {e}','danger')
            else:
                flash('Please select a valid Excel file.','warning')

        elif action=='clear_master':
            if request.form.get('confirm_clear')=='YES':
                db.execute("DELETE FROM branches")
                db.commit()
                flash('Branch master cleared.','success')
            else:
                flash('Type YES to confirm clear.','warning')

        elif action=='add_admin':
            u=request.form.get('username','').strip().lower()
            fn=request.form.get('full_name','').strip()
            p=request.form.get('password','')
            if not u or len(p)<6:
                flash('Username and password (min 6 chars) required.','warning')
            elif db.execute("SELECT 1 FROM admins WHERE username=?",(u,)).fetchone():
                flash(f'Username "{u}" already exists.','warning')
            else:
                db.execute("INSERT INTO admins(username,password_hash,full_name) VALUES(?,?,?)",
                           (u,generate_password_hash(p),fn)); db.commit()
                flash(f'OK Admin "{u}" created.','success')

        elif action=='del_admin':
            aid=request.form.get('admin_id')
            if aid and int(aid)!=1:
                db.execute("DELETE FROM admins WHERE id=?",(aid,)); db.commit()
                flash('Admin removed.','success')

        elif action=='reset_branch':
            bid=request.form.get('branch_id')
            if bid:
                db.execute("UPDATE branches SET password_hash=NULL,first_login=1 WHERE id=?",(bid,))
                db.commit(); flash('Password reset.','success')

        elif action=='del_branch':
            bid=request.form.get('branch_id')
            if bid:
                db.execute("DELETE FROM branches WHERE id=?",(bid,))
                db.commit(); flash('Branch removed.','success')

    admins   = db.execute("SELECT * FROM admins ORDER BY id").fetchall()
    branches = db.execute("SELECT * FROM branches ORDER BY zo,ro,branch_name").fetchall()
    zo_list  = sorted(set(b['zo'] for b in branches if b['zo']))
    db.close()
    return render_template_string(T_SETTINGS,
        admins=admins, branches=branches, zo_list=zo_list, week_now=week_now())

# ══════════════════════════════════════════════════════════════════════════════
# REMINDERS
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/reminders')
@admin_only
def reminders():
    week = request.args.get('week','')
    db   = db_conn()
    weeks = [r[0] for r in db.execute(
        "SELECT DISTINCT week_label FROM uploads ORDER BY uploaded_at DESC").fetchall()]
    if not week and weeks: week = weeks[0]
    pending=[]
    if week:
        rows = db.execute("""
            SELECT l.branch_code,l.branch_name,l.zo,l.ro,
                   COUNT(*) as total,
                   SUM(CASE WHEN l.status='Submitted' THEN 1 ELSE 0 END) as submitted,
                   b.email
            FROM loan_accounts l LEFT JOIN branches b ON l.branch_code=b.branch_code
            WHERE l.week_label=? GROUP BY l.branch_code
            HAVING submitted<total ORDER BY l.zo,l.ro
        """,(week,)).fetchall()
        pending = [dict(r) for r in rows]
        for b in pending:
            b['pct'] = round(b['submitted']/b['total']*100) if b['total'] else 0
    db.close()
    return render_template_string(T_REMINDERS,
        week=week,weeks=weeks,pending=pending,week_now=week_now())

# ══════════════════════════════════════════════════════════════════════════════
# BRANCH — LOGIN
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/branch', methods=['GET','POST'])
def branch_login():
    step = session.get('br_step','code')
    bc   = session.get('br_bc','')
    bn   = session.get('br_bn','')

    if request.method=='POST':
        act = request.form.get('action')

        if act=='check_code':
            code = request.form.get('branch_code','').strip().upper()
            db   = db_conn()
            row  = db.execute("SELECT * FROM branches WHERE branch_code=?",(code,)).fetchone()
            db.close()
            if not row:
                flash('Branch code not found. Please check and try again.','danger')
            else:
                session['br_bc']=code; session['br_bn']=row['branch_name'] or code
                session['br_step'] = 'set_pw' if (row['first_login'] or not row['password_hash']) else 'pw'
            return redirect(url_for('branch_login'))

        elif act=='set_pw':
            p1=request.form.get('password',''); p2=request.form.get('confirm','')
            if len(p1)<6: flash('Password must be at least 6 characters.','danger')
            elif p1!=p2:  flash('Passwords do not match.','danger')
            else:
                db=db_conn()
                db.execute("UPDATE branches SET password_hash=?,first_login=0 WHERE branch_code=?",
                           (generate_password_hash(p1),bc)); db.commit(); db.close()
                return _br_login_complete(bc)
            return redirect(url_for('branch_login'))

        elif act=='pw':
            p=request.form.get('password','')
            db=db_conn()
            row=db.execute("SELECT * FROM branches WHERE branch_code=?",(bc,)).fetchone()
            db.close()
            if row and check_password_hash(row['password_hash'],p):
                return _br_login_complete(bc)
            flash('Incorrect password.','danger')
            return redirect(url_for('branch_login'))

        elif act=='forgot':
            session['br_step']='forgot'
            return redirect(url_for('branch_login'))

        elif act=='verify_forgot':
            ans=request.form.get('answer','').strip().upper()
            if ans=='CGTMSE': session['br_step']='reset_pw'
            else: flash('Incorrect answer. Hint: name of this portal.','danger')
            return redirect(url_for('branch_login'))

        elif act=='reset_pw':
            p1=request.form.get('password',''); p2=request.form.get('confirm','')
            if len(p1)<6: flash('Password must be at least 6 characters.','danger')
            elif p1!=p2:  flash('Passwords do not match.','danger')
            else:
                db=db_conn()
                db.execute("UPDATE branches SET password_hash=?,first_login=0 WHERE branch_code=?",
                           (generate_password_hash(p1),bc)); db.commit(); db.close()
                return _br_login_complete(bc)
            return redirect(url_for('branch_login'))

    return render_template_string(T_BR_LOGIN, step=step, bc=bc, bn=bn)

def _br_login_complete(bc):
    db=db_conn()
    b=db.execute("SELECT * FROM branches WHERE branch_code=?",(bc,)).fetchone()
    db.close()
    session['br_step']='code'
    session.update({'role':'branch','br_code':bc,
                    'br_name':b['branch_name'] if b else bc,
                    'br_ro':b['ro'] if b else '',
                    'br_zo':b['zo'] if b else ''})
    return redirect(url_for('branch_home'))

@app.route('/branch/logout')
def branch_logout():
    session.clear()
    return redirect(url_for('branch_login'))

# ══════════════════════════════════════════════════════════════════════════════
# BRANCH HOME — sheet-wise tabs
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/branch/home')
@branch_only
def branch_home():
    bc   = session['br_code']
    week = request.args.get('week','')
    db   = db_conn()

    # All weeks that have accounts for this branch
    weeks = [r[0] for r in db.execute("""
        SELECT DISTINCT week_label FROM loan_accounts
        WHERE branch_code=? ORDER BY week_label DESC
    """,(bc,)).fetchall()]

    if not week and weeks: week = weeks[0]

    # Per-week stats (for tabs)
    week_stats = {}
    for w in weeks:
        tot = db.execute("SELECT COUNT(*) FROM loan_accounts WHERE branch_code=? AND week_label=?",(bc,w)).fetchone()[0]
        sub = db.execute("SELECT COUNT(*) FROM loan_accounts WHERE branch_code=? AND week_label=? AND status='Submitted'",(bc,w)).fetchone()[0]
        week_stats[w] = {'total':tot,'submitted':sub,
                         'pct':round(sub/tot*100) if tot else 0}

    accounts = []
    stats    = {'total':0,'submitted':0,'pending':0,'draft':0}
    if week:
        rows = db.execute("""
            SELECT l.id,l.status,l.row_data,l.branch_data,l.submitted_at,
                   u.col_config
            FROM loan_accounts l JOIN uploads u ON l.upload_id=u.id
            WHERE l.branch_code=? AND l.week_label=? ORDER BY l.id
        """,(bc,week)).fetchall()
        for r in rows:
            rd  = json.loads(r['row_data']   or '{}')
            bd  = json.loads(r['branch_data']or '{}')
            cfg = json.loads(r['col_config'] or '[]')
            # Build summary line from locked data (first few non-system filled fields)
            summary_fields = [(c['name'], rd.get(c['name'],''))
                              for c in cfg if c['type']=='locked'
                              and c['known'] not in ('zo','ro','branch_code','branch_name')
                              and rd.get(c['name'],'')][:3]
            accounts.append({'id':r['id'],'status':r['status'],
                             'summary':summary_fields,'submitted_at':r['submitted_at']})
            stats['total'] += 1
            s = (r['status'] or 'Pending').lower()
            stats[s] = stats.get(s,0)+1

    db.close()
    pct = round(stats['submitted']/stats['total']*100) if stats['total'] else 0
    return render_template_string(T_BR_HOME,
        bc=bc, bn=session['br_name'], ro=session.get('br_ro',''),
        week=week, weeks=weeks, week_stats=week_stats,
        accounts=accounts, stats=stats, pct=pct)

# ══════════════════════════════════════════════════════════════════════════════
# BRANCH ACCOUNT FORM — exact Excel columns
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/branch/account/<int:aid>', methods=['GET','POST'])
@branch_only
def branch_account(aid):
    bc = session['br_code']
    db = db_conn()
    acc = db.execute("""
        SELECT l.*,u.col_config FROM loan_accounts l
        JOIN uploads u ON l.upload_id=u.id
        WHERE l.id=? AND l.branch_code=?
    """,(aid,bc)).fetchone()
    if not acc: flash('Account not found.','danger'); return redirect(url_for('branch_home'))

    cfg = json.loads(acc['col_config'] or '[]')
    rd  = json.loads(acc['row_data']   or '{}')
    bd  = json.loads(acc['branch_data']or '{}')
    pf  = json.loads(acc['pdf_files']  or '{}')

    if request.method=='POST':
        action = request.form.get('_action','submit')

        # Collect fillable field values
        new_bd = dict(bd)
        for cc in cfg:
            if cc['type'] in ('fillable',):
                val = request.form.get(f"field_{cc['idx']}",'').strip()
                new_bd[cc['name']] = val

        # Handle PDF uploads
        new_pf = dict(pf)
        for cc in cfg:
            if cc['type']=='pdf':
                file_key = f"pdf_{cc['idx']}"
                uploaded = request.files.get(file_key)
                if uploaded and uploaded.filename:
                    fname = f"{bc}_{aid}_{cc['idx']}_{secure_filename(uploaded.filename)}"
                    uploaded.save(os.path.join(PDFS, fname))
                    new_pf[cc['name']] = fname

        new_status = 'Submitted' if action=='submit' else 'Draft'

        # Validate on submit: required + format check
        if action=='submit':
            errors = []
            import re as _re
            for cc in cfg:
                if cc['type'] != 'fillable': continue
                val = new_bd.get(cc['name'],'').strip()
                # Required check
                if cc.get('required') and not val:
                    errors.append(f'"{cc["name"]}" is required')
                    continue
                if not val: continue
                # Format validation
                fmt = cc.get('format','text')
                if fmt == 'number':
                    cleaned = val.replace(',','').replace(' ','')
                    if not _re.match(r'^-?\d+(\.\d+)?$', cleaned):
                        errors.append(f'"{cc["name"]}" must be a number (got: {val})')
                elif fmt == 'alphanumeric':
                    if not _re.match(r'^[A-Za-z0-9\s\-/]+$', val):
                        errors.append(f'"{cc["name"]}" must be letters/numbers only (got: {val})')
                elif fmt == 'dropdown':
                    opts = cc.get('options', [])
                    if opts and val not in opts:
                        errors.append(f'"{cc["name"]}" must be one of: {", ".join(opts)}')
            if errors:
                flash('Please fix: ' + ' | '.join(errors), 'warning')
                db.close()
                return redirect(url_for('branch_account', aid=aid))

        db.execute("""UPDATE loan_accounts SET branch_data=?,pdf_files=?,
                      status=?,submitted_at=? WHERE id=?""",
                   (json.dumps(new_bd), json.dumps(new_pf),
                    new_status,
                    datetime.now().strftime('%Y-%m-%d %H:%M') if action=='submit' else acc['submitted_at'],
                    aid))
        db.commit()

        if action=='submit':
            # Count remaining pending for flash message
            remaining = db.execute(
                "SELECT COUNT(*) FROM loan_accounts "
                "WHERE branch_code=? AND week_label=? AND status!='Submitted' AND id!=?",
                (bc, acc['week_label'], aid)).fetchone()[0]
            db.close()
            if remaining:
                flash(f'OK Account submitted! {remaining} account(s) still pending.', 'success')
            else:
                flash('🎉 All accounts submitted for this datasheet!', 'success')
            # Always go back to the sheet (home) page — branch decides what to do next
            return redirect(url_for('branch_home', week=acc['week_label']))
        db.close()
        flash('Draft saved.','info')
        return redirect(url_for('branch_account',aid=aid))

    # navigation
    all_ids = [r[0] for r in db.execute(
        "SELECT id FROM loan_accounts WHERE branch_code=? AND week_label=? ORDER BY id",
        (bc,acc['week_label'])).fetchall()]
    idx   = all_ids.index(aid) if aid in all_ids else 0
    total = len(all_ids)
    prev_id = all_ids[idx-1] if idx>0 else None
    next_id = all_ids[idx+1] if idx<total-1 else None
    sub  = sum(1 for i in all_ids
               if db.execute("SELECT status FROM loan_accounts WHERE id=?",(i,)).fetchone()['status']=='Submitted')
    pct  = round(sub/total*100) if total else 0

    db.close()
    return render_template_string(T_BR_ACCOUNT,
        acc=acc, cfg=cfg, rd=rd, bd=bd, pf=pf,
        bc=bc, bn=session['br_name'],
        idx=idx+1, total=total, pct=pct,
        prev_id=prev_id, next_id=next_id, sub=sub)

# ── Branch downloads their own sheet ─────────────────────────────────────────
@app.route('/branch/download')
@branch_only
def branch_download():
    bc   = session['br_code']
    week = request.args.get('week','')
    db   = db_conn()
    rows = db.execute("""
        SELECT l.*,u.col_config FROM loan_accounts l
        JOIN uploads u ON l.upload_id=u.id
        WHERE l.branch_code=? AND l.week_label=? ORDER BY l.id
    """,(bc,week)).fetchall()
    db.close()
    if not rows:
        flash('No data found.','warning'); return redirect(url_for('branch_home'))

    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = 'My Accounts'
    H_FILL  = PatternFill("solid",fgColor="0F172A")
    H_FONT  = Font(bold=True,color="FFFFFF",size=11)
    LK_FILL = PatternFill("solid",fgColor="F8FAFC")
    LK_FONT = Font(color="1E293B")
    FL_FILL = PatternFill("solid",fgColor="FFF9C4")
    FL_FONT = Font(color="1A202C")
    PD_FILL = PatternFill("solid",fgColor="E8F5E9")

    cfg = json.loads(rows[0]['col_config'] or '[]')
    headers = [c['name'] for c in cfg] + ['Status','Submitted At']
    ws.append(headers)
    for cell in ws[1]: cell.fill=H_FILL; cell.font=H_FONT
    ws.freeze_panes='A2'

    for row in rows:
        rd  = json.loads(row['row_data']   or '{}')
        bd  = json.loads(row['branch_data']or '{}')
        data_row = []
        for cc in cfg:
            if cc['type']=='locked':
                val = rd.get(cc['name'],'')
            elif cc['type']=='fillable':
                val = bd.get(cc['name'],'') or rd.get(cc['name'],'')
            else:
                val = 'PDF Uploaded' if json.loads(row['pdf_files'] or '{}').get(cc['name']) else '—'
            data_row.append(val)
        data_row += [row['status'], row['submitted_at'] or '']
        ws.append(data_row)

    # Style cells by type
    for row_cells in ws.iter_rows(min_row=2):
        for i, cell in enumerate(row_cells):
            if i < len(cfg):
                t = cfg[i]['type']
                if t=='locked':   cell.fill=LK_FILL; cell.font=LK_FONT
                elif t=='fillable': cell.fill=FL_FILL; cell.font=FL_FONT
                elif t=='pdf':    cell.fill=PD_FILL

    for i, col in enumerate(ws.columns,1):
        ws.column_dimensions[get_column_letter(i)].width = 22

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"DataSheet_{bc}_{week.replace(' ','_').replace('·','')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── Serve uploaded PDFs ───────────────────────────────────────────────────────
@app.route('/pdfs/<path:fname>')
@admin_only
def serve_pdf(fname):
    return send_from_directory(PDFS, fname)

# ══════════════════════════════════════════════════════════════════════════════
# TEMPLATES — shared CSS
# ══════════════════════════════════════════════════════════════════════════════
CSS = """
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700;800&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
/* ═══════════════════════════════════════════════
   MORNING SKY  —  Data Collection Portal
   Warm white · Soft sky sidebar · Pastel accents
═══════════════════════════════════════════════ */
*{box-sizing:border-box;margin:0;padding:0}
:root{
  /* Page backgrounds — warm white, barely tinted */
  --bg:        #f7f9fc;
  --bg2:       #f0f4f9;
  --surface:   #ffffff;
  --surface2:  #f5f7fa;
  --surface3:  #edf1f7;

  /* Borders — very soft */
  --border:    #e4e9f2;
  --border2:   #d1d9e8;

  /* Text — rich but not harsh */
  --text:      #18243d;
  --text2:     #475872;
  --text3:     #8898b4;

  /* Sidebar — sky blue, not navy */
  --sb-bg:     #2e7de9;
  --sb-bg2:    #1a6dd4;
  --sb-accent: #60aff8;
  --sb-text:   rgba(255,255,255,.92);
  --sb-muted:  rgba(255,255,255,.55);
  --sb-border: rgba(255,255,255,.14);

  /* Accent palette — vivid but warm */
  --sky:       #0ea5e9;
  --sky2:      #0284c7;
  --sky-lt:    #e0f2fe;
  --sky-md:    #bae6fd;

  --indigo:    #4f46e5;
  --indigo2:   #4338ca;
  --indigo-lt: #eef2ff;
  --indigo-md: #c7d2fe;

  --teal:      #0d9488;
  --teal2:     #0f766e;
  --teal-lt:   #ccfbf1;
  --teal-md:   #99f6e4;

  --amber:     #d97706;
  --amber2:    #b45309;
  --amber-lt:  #fef3c7;
  --amber-md:  #fde68a;

  --rose:      #e11d48;
  --rose2:     #be123c;
  --rose-lt:   #fff1f2;
  --rose-md:   #fecdd3;

  --violet:    #7c3aed;
  --violet2:   #6d28d9;
  --violet-lt: #f5f3ff;
  --violet-md: #ddd6fe;

  /* Shadows — crisp and airy */
  --sh-sm: 0 1px 3px rgba(24,36,61,.07), 0 1px 2px rgba(24,36,61,.04);
  --sh-md: 0 4px 16px rgba(24,36,61,.09), 0 2px 4px rgba(24,36,61,.05);
  --sh-lg: 0 12px 36px rgba(24,36,61,.12), 0 4px 8px rgba(24,36,61,.06);
}

html{scroll-behavior:smooth}
body{
  font-family:'Plus Jakarta Sans',sans-serif;
  background:var(--bg);
  color:var(--text);
  min-height:100vh;
  font-size:14px;
  line-height:1.55;
  -webkit-font-smoothing:antialiased;
}

/* Extremely subtle warm sky tint — barely visible, just alive */
body::before{
  content:'';position:fixed;inset:0;z-index:0;pointer-events:none;
  background:
    radial-gradient(ellipse 60% 40% at 0% 0%,   rgba(14,165,233,.05) 0%, transparent 60%),
    radial-gradient(ellipse 50% 35% at 100% 100%,rgba(79,70,229,.04) 0%, transparent 60%),
    radial-gradient(ellipse 40% 30% at 50% 60%,  rgba(13,148,136,.03) 0%, transparent 50%);
}

a{color:var(--sky2);text-decoration:none;transition:color .15s}
a:hover{color:var(--indigo)}

/* ── Layout ── */
.layout{display:flex;min-height:100vh;position:relative;z-index:1}
.sidebar{
  width:232px;
  background:linear-gradient(175deg, var(--sb-bg) 0%, var(--sb-bg2) 100%);
  display:flex;flex-direction:column;
  position:fixed;top:0;left:0;bottom:0;z-index:50;
  box-shadow:3px 0 20px rgba(46,125,233,.18);
}
.main{margin-left:232px;flex:1;display:flex;flex-direction:column;min-height:100vh;position:relative;z-index:1}
.topbar{
  background:rgba(255,255,255,.92);
  border-bottom:1px solid var(--border);
  padding:12px 22px;
  display:flex;align-items:center;justify-content:space-between;
  position:sticky;top:0;z-index:40;
  backdrop-filter:blur(10px);
  gap:12px;flex-wrap:wrap;
  box-shadow:var(--sh-sm);
}
.page{padding:22px;flex:1}

/* ── Sidebar ── */
.sb-logo{padding:20px 16px 14px;border-bottom:1px solid var(--sb-border)}
.sb-logo h1{
  font-size:13px;font-weight:800;color:#fff;letter-spacing:.1px;
  display:flex;align-items:center;gap:8px;
}
.sb-logo p{font-size:10px;color:var(--sb-muted);margin-top:3px;padding-left:24px}
.sb-week{
  margin:10px 10px 4px;
  background:rgba(255,255,255,.12);border:1px solid var(--sb-border);
  border-radius:8px;padding:7px 11px;font-size:10px;
  color:rgba(255,255,255,.75);font-weight:500;
}
.sb-sec{
  padding:14px 14px 4px;font-size:9px;font-weight:700;
  color:var(--sb-muted);text-transform:uppercase;letter-spacing:1.2px;
}
.sb-item{
  display:flex;align-items:center;gap:10px;
  padding:9px 12px;margin:2px 8px;border-radius:9px;
  cursor:pointer;color:var(--sb-text);font-size:12.5px;font-weight:500;
  transition:all .16s;text-decoration:none;
}
.sb-item:hover{background:rgba(255,255,255,.16);color:#fff;text-decoration:none}
.sb-item.active{
  background:rgba(255,255,255,.22);color:#fff;font-weight:700;
  border-left:3px solid rgba(255,255,255,.9);padding-left:9px;
}
.sb-icon{font-size:15px;width:20px;text-align:center;flex-shrink:0}
.sb-bottom{margin-top:auto;padding:10px 8px;border-top:1px solid var(--sb-border)}
.sb-user{display:flex;align-items:center;gap:9px;padding:8px 10px}
.su-av{
  width:32px;height:32px;border-radius:9px;
  background:rgba(255,255,255,.25);
  display:flex;align-items:center;justify-content:center;
  font-size:14px;font-weight:800;color:#fff;flex-shrink:0;
}
.su-name{font-size:12px;font-weight:700;color:#fff}
.su-role{font-size:9.5px;color:var(--sb-muted);margin-top:1px}
.su-out{margin-left:auto;background:none;border:none;cursor:pointer;
        color:var(--sb-muted);font-size:15px;padding:4px;transition:color .15s;border-radius:5px}
.su-out:hover{color:#fff;background:rgba(255,255,255,.12)}

/* ── Cards ── */
.card{
  background:var(--surface);border:1px solid var(--border);
  border-radius:14px;box-shadow:var(--sh-sm);
  margin-bottom:18px;overflow:hidden;
  transition:box-shadow .2s,border-color .2s;
}
.card:hover{box-shadow:var(--sh-md);border-color:var(--border2)}
.card-head{
  padding:14px 18px;border-bottom:1px solid var(--border);
  display:flex;align-items:center;gap:11px;background:var(--surface2);
}
.card-icon{width:36px;height:36px;border-radius:10px;
           display:flex;align-items:center;justify-content:center;font-size:16px;flex-shrink:0}
.ci-c{background:var(--sky-lt)}   .ci-v{background:var(--indigo-lt)}
.ci-e{background:var(--teal-lt)}  .ci-a{background:var(--amber-lt)}
.ci-r{background:var(--rose-lt)}  .ci-b{background:var(--sky-lt)}
.ci-gr{background:var(--surface3)}
.card-head h2{font-size:14px;font-weight:700;color:var(--text)}
.card-head p{font-size:11px;color:var(--text3);margin-top:1px}
.card-head-r{margin-left:auto;display:flex;gap:7px;align-items:center;flex-wrap:wrap}
.card-body{padding:18px}

/* ── KPI Stats ── */
.stats-row{display:grid;grid-template-columns:repeat(6,1fr);gap:12px;margin-bottom:18px}
.stat{
  background:var(--surface);border:1px solid var(--border);
  border-radius:14px;padding:16px;box-shadow:var(--sh-sm);
  transition:all .2s;position:relative;overflow:hidden;
}
.stat::before{
  content:'';position:absolute;top:0;left:0;right:0;height:3px;border-radius:3px 3px 0 0;
}
.stat:nth-child(1)::before{background:linear-gradient(90deg,#38bdf8,#0ea5e9)}
.stat:nth-child(2)::before{background:linear-gradient(90deg,#14b8a6,#0d9488)}
.stat:nth-child(3)::before{background:linear-gradient(90deg,#818cf8,#4f46e5)}
.stat:nth-child(4)::before{background:linear-gradient(90deg,#60a5fa,#2563eb)}
.stat:nth-child(5)::before{background:linear-gradient(90deg,#34d399,#059669)}
.stat:nth-child(6)::before{background:linear-gradient(90deg,#fbbf24,#d97706)}
.stat:hover{box-shadow:var(--sh-md);transform:translateY(-2px)}
.stat-lbl{font-size:9.5px;font-weight:700;color:var(--text3);
          text-transform:uppercase;letter-spacing:.7px;margin-bottom:8px}
.stat-val{font-size:26px;font-weight:800;line-height:1;color:var(--text)}
.stat-sub{font-size:10px;color:var(--text3);margin-top:4px}
.sv-c{color:var(--sky2)}   .sv-e{color:var(--teal)}   .sv-a{color:var(--amber)}
.sv-r{color:var(--rose)}   .sv-v{color:var(--indigo)} .sv-b{color:#2563eb}

/* ── Buttons ── */
.btn{
  display:inline-flex;align-items:center;gap:6px;
  font-family:'Plus Jakarta Sans',sans-serif;font-size:12.5px;font-weight:600;
  padding:8px 16px;border-radius:9px;border:none;cursor:pointer;
  transition:all .16s;text-decoration:none;white-space:nowrap;
}
.btn:hover{text-decoration:none;transform:translateY(-1px)}
.btn-primary{background:linear-gradient(135deg,var(--sky2),var(--indigo2));color:#fff;
             box-shadow:0 3px 10px rgba(14,165,233,.25)}
.btn-primary:hover{box-shadow:0 5px 16px rgba(14,165,233,.38)}
.btn-success{background:linear-gradient(135deg,var(--teal2),#065f46);color:#fff;
             box-shadow:0 3px 10px rgba(13,148,136,.22)}
.btn-success:hover{box-shadow:0 5px 16px rgba(13,148,136,.33)}
.btn-danger{background:var(--rose-lt);color:var(--rose2);border:1px solid var(--rose-md)}
.btn-danger:hover{background:var(--rose);color:#fff;box-shadow:0 3px 10px rgba(225,29,72,.25)}
.btn-outline{background:var(--surface);border:1.5px solid var(--border2);color:var(--text2)}
.btn-outline:hover{border-color:var(--sky2);color:var(--sky2);background:var(--sky-lt)}
.btn-amber{background:var(--amber-lt);color:var(--amber2);border:1px solid var(--amber-md)}
.btn-amber:hover{background:var(--amber);color:#fff}
.btn-sm{font-size:11.5px;padding:6px 12px;border-radius:7px}
.btn-xs{font-size:10.5px;padding:3px 9px;border-radius:6px}

/* ── Forms ── */
.fg{margin-bottom:14px}
.fl{display:block;font-size:11px;font-weight:600;color:var(--text2);
    margin-bottom:5px;letter-spacing:.2px;text-transform:uppercase}
.fi{
  width:100%;background:var(--surface);border:1.5px solid var(--border2);
  border-radius:9px;padding:10px 13px;
  font-family:'Plus Jakarta Sans',sans-serif;font-size:13px;
  color:var(--text);outline:none;transition:all .16s;
}
.fi:focus{border-color:var(--sky2);box-shadow:0 0 0 3px rgba(14,165,233,.1)}
select.fi{cursor:pointer}

/* ── Table ── */
.tbl-wrap{overflow-x:auto}
table{width:100%;border-collapse:collapse;font-size:12.5px}
thead tr{background:var(--surface2)}
th{padding:10px 14px;text-align:left;font-size:9.5px;font-weight:700;
   color:var(--text3);text-transform:uppercase;letter-spacing:.6px;
   border-bottom:1px solid var(--border);white-space:nowrap}
tbody tr{border-bottom:1px solid var(--border);transition:background .12s}
tbody tr:last-child{border-bottom:none}
tbody tr:hover{background:#f0f7ff}
td{padding:10px 14px;vertical-align:middle}

/* ── Badges ── */
.badge{display:inline-flex;align-items:center;gap:3px;font-size:10.5px;
       font-weight:700;padding:3px 9px;border-radius:100px;white-space:nowrap}
.b-e{background:var(--teal-lt);color:var(--teal2)}
.b-a{background:var(--amber-lt);color:var(--amber2)}
.b-r{background:var(--rose-lt);color:var(--rose2)}
.b-c{background:var(--sky-lt);color:var(--sky2)}
.b-v{background:var(--violet-lt);color:var(--violet2)}
.b-gr{background:var(--surface3);color:var(--text3)}
.bc-tag{font-family:'DM Mono',monospace;font-size:10.5px;font-weight:500;
        color:var(--sky2);background:var(--sky-lt);
        padding:2px 8px;border-radius:5px;border:1px solid var(--sky-md)}

/* ── Progress ── */
.prog{display:flex;align-items:center;gap:7px}
.prog-bar{flex:1;height:5px;background:var(--surface3);border-radius:3px;overflow:hidden;min-width:50px}
.prog-fill{height:100%;border-radius:3px;transition:width .4s ease;
           background:linear-gradient(90deg,var(--sky2),var(--indigo2))}
.prog-pct{font-size:10.5px;font-weight:700;color:var(--text3);
          width:28px;font-family:'DM Mono',monospace}

/* ── Upload zone ── */
.uz{border:2px dashed var(--border2);border-radius:13px;padding:34px 20px;
    text-align:center;cursor:pointer;transition:all .2s;background:var(--surface2);position:relative}
.uz:hover,.uz.drag{border-color:var(--sky2);background:var(--sky-lt);
                   box-shadow:0 0 0 4px rgba(14,165,233,.06)}
.uz input[type=file]{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%}

/* ── Alerts ── */
.alert{border-radius:10px;padding:11px 15px;margin-bottom:14px;
       font-size:12.5px;font-weight:500;
       display:flex;align-items:flex-start;gap:8px;line-height:1.5}
.al-s{background:var(--teal-lt);color:var(--teal2);border:1px solid var(--teal-md)}
.al-d{background:var(--rose-lt);color:var(--rose2);border:1px solid var(--rose-md)}
.al-w{background:var(--amber-lt);color:var(--amber2);border:1px solid var(--amber-md)}
.al-i{background:var(--sky-lt);color:var(--sky2);border:1px solid var(--sky-md)}

/* ── Dashboard drill-down tree ── */
.zo-block{border:1px solid var(--border);border-radius:13px;overflow:hidden;
          margin-bottom:10px;background:var(--surface);box-shadow:var(--sh-sm);transition:box-shadow .18s}
.zo-block:hover{box-shadow:var(--sh-md)}
.zo-head{
  padding:13px 16px;cursor:pointer;
  display:flex;align-items:center;gap:10px;
  font-weight:700;font-size:13px;color:var(--text);
  border-bottom:1px solid transparent;
  transition:background .15s;
  background:linear-gradient(90deg,#eef6ff,var(--surface));
}
.zo-head:hover{background:var(--sky-lt)}
.zo-head.open{border-bottom-color:var(--border);background:#eef6ff}
.zo-body{display:none;padding:12px;background:var(--surface2)}
.ro-block{margin-bottom:8px;border:1px solid var(--border);border-radius:10px;overflow:hidden;background:var(--surface)}
.ro-head{background:var(--surface2);padding:9px 14px;cursor:pointer;
         display:flex;align-items:center;gap:8px;
         font-weight:600;font-size:12px;color:var(--text);transition:background .14s}
.ro-head:hover{background:var(--sky-lt)}
.ro-body{display:none;padding:8px;background:var(--surface)}
.br-row{
  display:flex;align-items:center;gap:10px;padding:9px 12px;
  border-radius:8px;border:1px solid var(--border);margin-bottom:5px;
  cursor:pointer;background:var(--surface2);transition:all .16s;
}
.br-row:hover{border-color:var(--sky2);background:#e0f2fe;
              box-shadow:0 2px 8px rgba(14,165,233,.1)}
.br-row:last-child{margin-bottom:0}

/* ── Modal ── */
.modal-bg{display:none;position:fixed;inset:0;background:rgba(24,36,61,.4);
          z-index:200;align-items:center;justify-content:center;padding:20px;
          backdrop-filter:blur(4px)}
.modal-bg.open{display:flex}
.modal-box{
  background:var(--surface);border:1px solid var(--border);
  border-radius:18px;width:100%;max-width:800px;max-height:88vh;
  overflow-y:auto;box-shadow:var(--sh-lg);
}
.modal-head{
  padding:16px 20px;border-bottom:1px solid var(--border);
  display:flex;align-items:center;justify-content:space-between;
  position:sticky;top:0;background:var(--surface);z-index:1;
}
.modal-close{background:none;border:none;font-size:18px;cursor:pointer;
             color:var(--text3);transition:all .14s;line-height:1;
             width:28px;height:28px;border-radius:6px;display:flex;align-items:center;justify-content:center}
.modal-close:hover{color:var(--rose2);background:var(--rose-lt)}

/* ── Mapping column cards ── */
.col-card{background:var(--surface2);border:1.5px solid var(--border);
          border-radius:10px;display:flex;align-items:center;
          gap:10px;margin-bottom:8px;transition:all .16s}
.col-name{font-weight:600;font-size:12.5px;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.col-type-sel{display:flex;gap:4px}
.type-btn{font-size:11px;font-weight:600;padding:4px 10px;border-radius:6px;cursor:pointer;
          border:1.5px solid var(--border2);background:var(--surface);color:var(--text2);
          font-family:'Plus Jakarta Sans',sans-serif;transition:all .14s}
.type-btn:hover{color:var(--text);border-color:var(--border2)}
.type-btn.selected-locked  {background:#1e293b;color:#fff;border-color:#1e293b}
.type-btn.selected-fillable{background:var(--amber-lt);color:var(--amber2);border-color:var(--amber-md)}
.type-btn.selected-pdf     {background:var(--violet-lt);color:var(--violet2);border-color:var(--violet-md)}

/* ── Field type tags ── */
.tag-locked  {background:#1e293b;color:#cbd5e1;font-size:9.5px;font-weight:700;padding:2px 7px;border-radius:4px}
.tag-fillable{background:var(--amber-lt);color:var(--amber2);font-size:9.5px;font-weight:700;padding:2px 7px;border-radius:4px}
.tag-pdf     {background:var(--violet-lt);color:var(--violet2);font-size:9.5px;font-weight:700;padding:2px 7px;border-radius:4px}
.tag-sys     {background:var(--sky-lt);color:var(--sky2);font-size:9.5px;font-weight:700;padding:2px 7px;border-radius:4px}

/* ── Branch portal ── */
.br-wrap{max-width:660px;margin:0 auto;padding:16px}
.br-hdr{
  background:linear-gradient(135deg,var(--sb-bg),var(--sb-bg2));
  padding:14px 20px;
  display:flex;align-items:center;justify-content:space-between;
  position:sticky;top:0;z-index:40;
  box-shadow:0 2px 12px rgba(46,125,233,.2);
}
.br-hdr-prog{height:3px;background:rgba(255,255,255,.2)}
.br-hdr-fill{
  height:100%;
  background:linear-gradient(90deg,#fff,rgba(255,255,255,.6),#fff);
  background-size:200% 100%;
  animation:shimmer 2.5s linear infinite;
  transition:width .4s ease;
}
@keyframes shimmer{0%{background-position:200%}100%{background-position:0%}}

/* ── Account cards (branch) ── */
.acc-card{
  background:var(--surface);border:1.5px solid var(--border);
  border-radius:12px;padding:14px 16px;margin-bottom:8px;
  cursor:pointer;transition:all .16s;
  display:flex;align-items:center;gap:12px;
  position:relative;overflow:hidden;
  text-decoration:none;color:var(--text);box-shadow:var(--sh-sm);
}
.acc-card::before{content:'';position:absolute;left:0;top:0;bottom:0;width:4px}
.acc-card.pend::before{background:var(--amber)}
.acc-card.subm::before{background:var(--teal)}
.acc-card.draft::before{background:var(--sky2)}
.acc-card:hover{border-color:var(--sky2);transform:translateY(-1px);
                box-shadow:var(--sh-md);text-decoration:none;color:var(--text)}

/* ── Datasheet tabs ── */
.ds-tabs{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:16px}
.ds-tab{padding:7px 14px;border-radius:8px;font-size:12px;font-weight:600;
        cursor:pointer;border:1.5px solid var(--border2);background:var(--surface);
        color:var(--text2);text-decoration:none;transition:all .16s}
.ds-tab:hover{border-color:var(--sky2);color:var(--sky2);background:var(--sky-lt);text-decoration:none}
.ds-tab.active{background:var(--sb-bg);border-color:var(--sb-bg);
               color:#fff;font-weight:700;box-shadow:0 3px 10px rgba(46,125,233,.25)}

/* ── Branch form fields ── */
.field-row{display:flex;align-items:flex-start;gap:12px;margin-bottom:13px;
           padding-bottom:13px;border-bottom:1px solid var(--border)}
.field-row:last-child{border-bottom:none;margin-bottom:0;padding-bottom:0}
.field-label{font-size:10.5px;font-weight:700;color:var(--text3);width:150px;
             flex-shrink:0;padding-top:10px;text-transform:uppercase;letter-spacing:.4px;line-height:1.6}
.field-locked{flex:1;font-size:13px;font-weight:500;color:var(--text);
              padding:9px 13px;background:var(--surface2);border-radius:8px;
              border:1.5px solid var(--border);min-height:38px;word-break:break-all}
.field-locked.empty{color:var(--text3);font-style:italic}
.field-fillable{flex:1}
.field-fillable input,.field-fillable textarea,.field-fillable select{
  width:100%;border:1.5px solid #fbbf24;border-radius:8px;padding:9px 13px;
  font-family:'Plus Jakarta Sans',sans-serif;font-size:13px;
  background:#fffbeb;color:var(--text);outline:none;transition:all .16s}
.field-fillable input:focus,.field-fillable textarea:focus,.field-fillable select:focus{
  border-color:var(--sky2);background:#f0f9ff;box-shadow:0 0 0 3px rgba(14,165,233,.1)}
.field-pdf{flex:1}
.field-pdf input[type=file]{border:1.5px solid #a5b4fc;border-radius:8px;padding:8px 12px;
                            width:100%;background:var(--indigo-lt);font-size:12px;cursor:pointer}
.pdf-done{display:flex;align-items:center;gap:7px;padding:8px 13px;
          background:var(--indigo-lt);border:1.5px solid var(--indigo-md);
          border-radius:8px;font-size:12px;color:var(--indigo2)}

/* ── Login ── */
.login-page{min-height:100vh;background:var(--bg);
            display:flex;align-items:center;justify-content:center;
            padding:20px;position:relative;z-index:1}
.login-card{background:var(--surface);border:1px solid var(--border);
            border-radius:22px;width:100%;max-width:400px;overflow:hidden;
            box-shadow:var(--sh-lg);animation:slideUp .42s ease}
@keyframes slideUp{from{opacity:0;transform:translateY(20px)}to{opacity:1;transform:translateY(0)}}
.login-hero{
  padding:32px 28px 26px;text-align:center;
  background:linear-gradient(135deg,var(--sb-bg) 0%,var(--sb-bg2) 100%);
  position:relative;overflow:hidden;
}
.login-hero::before{
  content:'';position:absolute;inset:0;
  background:radial-gradient(ellipse at 30% 50%,rgba(255,255,255,.12),transparent 55%),
             radial-gradient(ellipse at 70% 40%,rgba(255,255,255,.08),transparent 55%);
}
.login-hero-icon{font-size:46px;margin-bottom:12px;position:relative;
                 filter:drop-shadow(0 4px 10px rgba(0,0,0,.2))}
.login-hero h1{font-size:20px;font-weight:800;color:#fff;margin-bottom:4px;position:relative}
.login-hero p{font-size:11px;color:rgba(255,255,255,.55);position:relative}
.login-body{padding:26px}
.pw-wrap{position:relative}
.pw-toggle{position:absolute;right:11px;top:50%;transform:translateY(-50%);
           background:none;border:none;cursor:pointer;font-size:14px;
           color:var(--text3);transition:color .14s}
.pw-toggle:hover{color:var(--sky2)}
.strength-bar{display:flex;gap:3px;margin-top:5px}
.s-seg{height:3px;flex:1;border-radius:2px;background:var(--surface3);transition:background .3s}
.s-seg.w{background:var(--rose)} .s-seg.f{background:var(--amber)} .s-seg.s{background:var(--teal)}

/* ── Empty states ── */
.empty{text-align:center;padding:50px 20px;color:var(--text3)}
.empty .ei{font-size:48px;opacity:.25;margin-bottom:12px;display:block}
.empty h3{font-size:15px;font-weight:700;color:var(--text2);margin-bottom:6px}
.empty p{font-size:12.5px}

/* ── Page entry animations ── */
.page>*{animation:fadeUp .3s ease both}
.page>*:nth-child(1){animation-delay:.04s}
.page>*:nth-child(2){animation-delay:.08s}
.page>*:nth-child(3){animation-delay:.12s}
.page>*:nth-child(4){animation-delay:.16s}
.page>*:nth-child(5){animation-delay:.20s}
@keyframes fadeUp{from{opacity:0;transform:translateY(9px)}to{opacity:1;transform:translateY(0)}}

/* ── Misc ── */
.topbar-title{font-size:15px;font-weight:700;color:var(--text)}
.topbar-sub{font-size:11px;color:var(--text3);margin-top:1px}
.link-btn{background:none;border:none;color:var(--sky2);cursor:pointer;
          font-size:12px;font-weight:600;font-family:inherit;padding:0;transition:color .14s}
.link-btn:hover{color:var(--indigo)}
.divider{display:flex;align-items:center;gap:10px;margin:14px 0;color:var(--text3);font-size:11px}
.divider::before,.divider::after{content:'';flex:1;height:1px;background:var(--border)}
.pill{display:flex;align-items:center;gap:10px;background:var(--surface2);
      border:1px solid var(--border);border-radius:9px;padding:10px 14px;transition:all .16s}
.pill:hover{border-color:var(--border2)}
::-webkit-scrollbar{width:5px;height:5px}
::-webkit-scrollbar-track{background:transparent}
::-webkit-scrollbar-thumb{background:var(--border2);border-radius:3px}
::-webkit-scrollbar-thumb:hover{background:var(--text3)}
@media(max-width:900px){
  .sidebar{display:none}.main{margin-left:0}
  .stats-row{grid-template-columns:repeat(3,1fr)}
}
@media(max-width:600px){.stats-row{grid-template-columns:repeat(2,1fr)}}
</style>"""

FLASH = """{% with msgs=get_flashed_messages(with_categories=true) %}
  {% for cat,msg in msgs %}
  <div class="alert {{'al-s' if cat=='success' else 'al-d' if cat=='danger' else 'al-w' if cat=='warning' else 'al-i'}}">
    <span>{{'OK' if cat=='success' else '❌' if cat=='danger' else '️' if cat=='warning' else 'ℹ️'}}</span>
    <span>{{msg}}</span>
  </div>
  {% endfor %}{% endwith %}"""


NAV = """<div class="sidebar">
  <div class="sb-logo">
    <h1> Data Collection Portal</h1>
    <p>Central Office · v3</p>
  </div>
  <div style="padding:8px 10px 4px">
    <div style="background:var(--surface);border:1px solid var(--border);border-radius:8px;
                padding:7px 11px;font-size:10px;color:var(--text2);font-weight:500">
      📅 {{week_now}}
    </div>
  </div>
  <div class="sb-sec">Main</div>
  <a href="/dashboard" class="sb-item {{'active' if active=='dashboard'}}">
    <span class="sb-icon">✦</span> Dashboard</a>
  <a href="/upload" class="sb-item {{'active' if active=='upload'}}">
    <span class="sb-icon">Up</span> Upload Datasheet</a>
  <div class="sb-sec">Actions</div>
  <a href="/reminders" class="sb-item {{'active' if active=='reminders'}}">
    <span class="sb-icon">◈</span> Reminders</a>
  <a href="/export" class="sb-item {{'active' if active=='export'}}">
    <span class="sb-icon">Dn</span> Export</a>
  <div class="sb-sec">Settings</div>
  <a href="/settings" class="sb-item {{'active' if active=='settings'}}">
    <span class="sb-icon">◉</span> Settings</a>
  <div class="sb-bottom">
    <div class="sb-user">
      <div class="su-av">{{session.full_name[0]|upper if session.full_name else 'A'}}</div>
      <div>
        <div class="su-name">{{session.full_name}}</div>
        <div class="su-role">Administrator</div>
      </div>
      <a href="/logout" class="su-out" title="Sign Out">⏻</a>
    </div>
  </div>
</div>"""


# ══════════════════════════════════════════════════════════════════════════════
# TEMPLATE: ADMIN LOGIN
# ══════════════════════════════════════════════════════════════════════════════
T_LOGIN = CSS + """
<div class="login-page">
  <div class="login-card">
    <div class="login-hero">
      <div class="login-hero-icon"></div>
      <h1>Data Collection Portal</h1>
      <p>Central Office · Admin Panel</p>
    </div>
    <div class="login-body">""" + FLASH + """
      <form method="POST">
        <div class="fg"><label class="fl">USERNAME</label>
          <input name="username" class="fi mono" type="text" placeholder="admin" required autofocus autocomplete="username"/>
        </div>
        <div class="fg"><label class="fl">PASSWORD</label>
          <div style="position:relative">
            <input name="password" class="fi" id="ap" type="password" placeholder="Password" required/>
            <button type="button" style="position:absolute;right:10px;top:50%;transform:translateY(-50%);
              background:none;border:none;cursor:pointer;font-size:15px;color:var(--text3)"
              onclick="const i=document.getElementById('ap');i.type=i.type==='password'?'text':'password'">👁️</button>
          </div>
        </div>
        <button class="btn btn-primary" style="width:100%;justify-content:center;padding:11px;font-size:14px">
          Sign In →
        </button>
      </form>
      <div style="text-align:center;margin-top:18px;padding-top:16px;border-top:1px solid var(--border)">
        <a href="/branch" style="font-size:12px;color:var(--text3)">Branch Login →</a>
      </div>
    </div>
  </div>
</div>"""


# ══════════════════════════════════════════════════════════════════════════════
# TEMPLATE: FIX ZONES
# ══════════════════════════════════════════════════════════════════════════════
T_FIX_ZONES = CSS + """<div class="layout">""" + NAV + """
<div class="main">
  <div class="topbar">
    <div>
      <div class="topbar-title">️ Assign Missing Zones (ZO)</div>
      <div class="topbar-sub">Optional — skip if not needed now</div>
    </div>
  </div>
  <div class="page">""" + FLASH + """
    <div class="card">
      <div class="card-head">
        <div class="card-icon ci-a">🗺️</div>
        <div>
          <h2>{{fname}} — Missing Zone (ZO)</h2>
          <p>{{missing|length}} region(s) have no zone assigned. Fill in or skip.</p>
        </div>
        <div class="card-head-r">
          <form method="POST">
            <button name="skip" value="1" class="btn btn-outline btn-sm">Skip for now →</button>
          </form>
        </div>
      </div>
      <div class="card-body">
        <div class="al-i alert" style="margin-bottom:16px">
          ℹ️ Zone (ZO) is optional. Branches can still log in and submit without it.
          You can always update zones later in <strong>Settings → Branch Master</strong>.
        </div>

        <!-- Datalist for autocomplete -->
        <datalist id="zoList">
          {% for z in all_zo %}<option value="{{z}}">{% endfor %}
        </datalist>

        <form method="POST">
          <div style="display:grid;gap:10px;margin-bottom:18px">
            {% for row in missing %}
            <div style="display:flex;align-items:center;gap:14px;padding:14px 16px;
                        background:var(--amber-lt);border:1.5px solid #fde68a;border-radius:10px">
              <div style="flex:1">
                <div style="font-size:13px;font-weight:700;margin-bottom:2px">📍 {{row.ro}}</div>
                <div style="font-size:11px;color:var(--amber)">
                  {{row.branches}} branch(es) · {{row.accounts}} accounts
                </div>
              </div>
              <div style="display:flex;align-items:center;gap:8px;flex-shrink:0">
                <label style="font-size:11px;font-weight:600;color:var(--text3)">Assign ZO:</label>
                <input type="text" name="zo_{{row.ro}}" list="zoList"
                       class="fi" placeholder="Type or select zone..."
                       style="width:220px;padding:8px 12px;font-size:13px"
                       autocomplete="off"/>
              </div>
            </div>
            {% endfor %}
          </div>
          <div style="display:flex;gap:10px;justify-content:flex-end">
            <button name="skip" value="1" class="btn btn-outline">Skip →</button>
            <button type="submit" class="btn btn-success">OK Save Zones & Continue</button>
          </div>
        </form>
      </div>
    </div>
  </div>
</div></div>
"""

# ══════════════════════════════════════════════════════════════════════════════
# TEMPLATE: UPLOAD (Step 1)
# ══════════════════════════════════════════════════════════════════════════════
T_UPLOAD = CSS + """<div class="layout">""" + NAV + """
<div class="main">
  <div class="topbar">
    <div><div class="topbar-title">Upload Datasheet</div>
         <div class="topbar-sub">System auto-detects columns → you confirm mapping</div></div>
  </div>
  <div class="page">""" + FLASH + """
    <div style="display:grid;grid-template-columns:1.4fr 1fr;gap:18px">
      <div>
        <div class="card">
          <div class="card-head"><div class="card-icon ci-c"></div>
            <div><h2>Upload Excel File</h2><p>Step 1 of 2 — upload, then review column mapping</p></div></div>
          <div class="card-body">
            <form method="POST" enctype="multipart/form-data" id="upForm">
              <div class="uz" id="dz">
                <input type="file" name="file" id="fi" accept=".xlsx,.xls" required/>
                <div style="font-size:36px;margin-bottom:10px" id="uzIcon"></div>
                <div style="font-size:14px;font-weight:700;margin-bottom:4px" id="uzTitle">Drop your Excel file here</div>
                <div style="font-size:11px;color:var(--text3)" id="uzSub">or click to browse · .xlsx/.xls</div>
              </div>
              <div style="display:flex;gap:10px;align-items:flex-end;margin-top:14px;flex-wrap:wrap">
                <div style="flex:1;min-width:180px">
                  <label class="fl">DATASHEET LABEL</label>
                  <input name="week_label" class="fi" type="text" value="{{week_now}}" placeholder="e.g. Sheet 10 · Mar 2026"/>
                </div>
                <button type="submit" class="btn btn-primary" id="upBtn">
                  Next: Map Fields →
                </button>
              </div>
            </form>
          </div>
        </div>

        <div class="card">
          <div class="card-head"><div class="card-icon ci-gr">🗂️</div>
            <div><h2>Uploaded Files</h2><p>Click datasheet to view on dashboard</p></div></div>
          <div class="tbl-wrap">{% if recent %}
            <table><thead><tr><th>File</th><th>Datasheet</th><th>Accounts</th><th>By</th><th>Date</th><th></th></tr></thead>
            <tbody>{% for u in recent %}
            <tr>
              <td style="max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-weight:500">
                 {{u.filename}}</td>
              <td><span class="badge b-c" style="font-size:10px">{{u.week_label}}</span></td>
              <td style="text-align:center;font-weight:600">{{u.accounts}}</td>
              <td style="color:var(--text3);font-size:11px">{{u.uploaded_by}}</td>
              <td style="color:var(--text3);font-size:11px">{{u.uploaded_at[:16]}}</td>
              <td>
                <div style="display:flex;gap:5px">
                  <a href="/dashboard?week={{u.week_label|urlencode}}" class="btn btn-xs btn-outline"></a>
                  <form method="POST" action="/upload/delete/{{u.id}}" style="display:inline"
                        onsubmit="return confirm('Delete upload and all its data?')">
                    <button class="btn btn-xs btn-danger"></button>
                  </form>
                </div>
              </td>
            </tr>{% endfor %}</tbody></table>
          {% else %}
          <div class="empty" style="padding:24px"><div class="ei"></div><p>No files uploaded yet.</p></div>
          {% endif %}</div>
        </div>
      </div>

      <!-- Guide -->
      <div>
        <div class="card">
          <div class="card-head"><div class="card-icon ci-e"></div>
            <div><h2>Column Guide</h2><p>System recognizes these automatically</p></div></div>
          <div class="card-body" style="padding:0">
            <table><thead><tr><th>Data</th><th>Recognized As</th><th>Required?</th></tr></thead>
            <tbody>
              <tr><td style="font-weight:600">ZO / Zone</td>
                  <td style="font-size:11px;color:var(--text3)">ZO, Zone, Zonal Office</td>
                  <td><span class="badge b-a">Optional</span></td></tr>
              <tr><td style="font-weight:600">RO / Region</td>
                  <td style="font-size:11px;color:var(--text3)">RO, Region, Regional Office</td>
                  <td><span class="badge b-r">Mandatory</span></td></tr>
              <tr><td style="font-weight:600">Branch Code</td>
                  <td style="font-size:11px;color:var(--text3)">Branch Code, Sol ID, Code</td>
                  <td><span class="badge b-a">Optional</span></td></tr>
              <tr><td style="font-weight:600">Branch Name</td>
                  <td style="font-size:11px;color:var(--text3)">Branch Name, Br Name</td>
                  <td><span class="badge b-a">Optional</span></td></tr>
              <tr><td style="font-weight:600">All other columns</td>
                  <td style="font-size:11px;color:var(--text3)">Shown as-is to branch</td>
                  <td><span class="badge b-gr">Auto-typed</span></td></tr>
            </tbody></table>
            <div style="padding:10px 14px;background:var(--amber-lt);border-top:1px solid #fde68a;
                        font-size:11px;color:var(--amber)">
              ⚡ Columns with data → <strong>Locked</strong> · Empty columns → <strong>Fillable by branch</strong><br>
              You can change any column type in Step 2 (Map Fields).
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>
</div></div>
<script>
const dz=document.getElementById('dz'),fi=document.getElementById('fi');
fi.addEventListener('change',e=>{if(e.target.files[0]){
  document.getElementById('uzIcon').textContent='📄';
  document.getElementById('uzTitle').textContent=e.target.files[0].name;
  document.getElementById('uzSub').textContent='Ready · click "Next: Map Fields"'}});
dz.addEventListener('dragover',e=>{e.preventDefault();dz.classList.add('drag')});
dz.addEventListener('dragleave',()=>dz.classList.remove('drag'));
dz.addEventListener('drop',e=>{e.preventDefault();dz.classList.remove('drag');
  const f=Array.from(e.dataTransfer.files).find(f=>/\.xlsx?$/i.test(f.name));
  if(f){const dt=new DataTransfer();dt.items.add(f);fi.files=dt.files;
    document.getElementById('uzIcon').textContent='📄';
    document.getElementById('uzTitle').textContent=f.name;
    document.getElementById('uzSub').textContent='Ready'}});
document.getElementById('upForm').addEventListener('submit',()=>{
  document.getElementById('upBtn').textContent=' Reading...';
  document.getElementById('upBtn').disabled=true;});
</script>"""

# ══════════════════════════════════════════════════════════════════════════════
# TEMPLATE: MAP FIELDS (Step 2)
# ══════════════════════════════════════════════════════════════════════════════
T_MAP_FIELDS = CSS + """<div class="layout">""" + NAV + """
<div class="main">
  <div class="topbar">
    <div>
      <div class="topbar-title">Map Fields</div>
      <div class="topbar-sub">Step 2 of 2 — set type &amp; format for each column, then save</div>
    </div>
    <div style="display:flex;gap:8px">
      <a href="/upload" class="btn btn-outline btn-sm">← Back</a>
    </div>
  </div>
  <div class="page">""" + FLASH + """
    <div style="display:grid;grid-template-columns:1fr 270px;gap:18px;align-items:start">
      <div>
        <div class="card">
          <div class="card-head">
            <div class="card-icon ci-v">🗂️</div>
            <div>
              <h2>{{pending.filename}}</h2>
              <p>{{pending.data|length}} rows · set type and format for each column</p>
            </div>
          </div>
          <div class="card-body">

            <!-- Legend -->
            <div style="display:flex;gap:6px;flex-wrap:wrap;margin-bottom:16px;
                        padding:10px 12px;background:var(--bg);border-radius:9px;align-items:center">
              <span style="font-size:10px;font-weight:700;color:var(--text3);text-transform:uppercase;letter-spacing:.5px;margin-right:4px">Types:</span>
              <span class="tag-locked">🔒 Locked</span>
              <span class="tag-fillable">✏️ Fillable</span>
              <span class="tag-pdf">📎 PDF</span>
              <span class="tag-sys">⚙️ System</span>
            </div>

            <form method="POST" id="mapForm">
              {% for cc in pending.col_cfg %}

              <!-- ── Column card ── -->
              <div class="col-card" id="card_{{cc.idx}}"
                   style="flex-direction:column;align-items:stretch;gap:0;padding:0;overflow:hidden;
                          border-color:{{'#2563eb' if cc.type=='fillable' else '#7c3aed' if cc.type=='pdf' else '#e2e8f0'}}">

                <!-- Top row: index + name + type buttons + preview -->
                <div style="display:flex;align-items:center;gap:10px;padding:11px 14px">
                  <div style="width:22px;height:22px;border-radius:5px;flex-shrink:0;
                              display:flex;align-items:center;justify-content:center;
                              font-size:10px;font-weight:800;
                              {{'background:#1e293b;color:#fff' if cc.known else
                                'background:var(--amber-lt);color:var(--amber)' if cc.type=='fillable' else
                                'background:var(--purple-lt);color:var(--purple)' if cc.type=='pdf' else
                                'background:#f1f5f9;color:var(--text3)'}}">{{loop.index}}</div>

                  <div style="flex:1;min-width:0">
                    <div class="col-name">{{cc.name}}</div>
                    <div style="font-size:9.5px;color:var(--text3);margin-top:1px">
                      Filled: {{(cc.fill_ratio*100)|int}}%
                      {% if cc.known %} · <span style="color:var(--blue);font-weight:600">System: {{cc.known}}</span>{% endif %}
                      &nbsp;·&nbsp; Preview:
                      {% for row in pending.data[:2] %}
                        <span style="font-family:monospace">{{ row[cc.idx] if cc.idx < row|length and row[cc.idx] else '—' }}</span>{% if not loop.last %},&nbsp;{% endif %}
                      {% endfor %}
                    </div>
                  </div>

                  <!-- Type buttons -->
                  {% if cc.known in ('zo','ro','branch_code','branch_name') %}
                    <span class="tag-sys">⚙️ SYSTEM</span>
                    <input type="hidden" name="col_{{cc.idx}}" value="locked"/>
                  {% else %}
                    <div class="col-type-sel" id="tsel_{{cc.idx}}">
                      {% for t,lbl in [('locked','🔒 Locked'),('fillable','✏️ Fillable'),('pdf','📎 PDF')] %}
                      <button type="button"
                        class="type-btn {{'selected-'+cc.type if cc.type==t else ''}}"
                        id="tb_{{cc.idx}}_{{t}}"
                        onclick="setType({{cc.idx}},'{{t}}')">{{lbl}}</button>
                      {% endfor %}
                      <input type="hidden" name="col_{{cc.idx}}" value="{{cc.type}}" id="hid_{{cc.idx}}"/>
                    </div>
                  {% endif %}
                </div>

                <!-- Fillable options panel — shown/hidden by JS -->
                {% if cc.known not in ('zo','ro','branch_code','branch_name') %}
                <div id="opts_{{cc.idx}}"
                     style="display:{{'block' if cc.type=='fillable' else 'none'}};
                            border-top:1px solid var(--border);background:#fffbeb;padding:10px 14px">
                  <div style="display:flex;gap:14px;flex-wrap:wrap;align-items:flex-end">

                    <!-- Format -->
                    <div>
                      <div style="font-size:9.5px;font-weight:700;color:var(--amber);text-transform:uppercase;
                                  letter-spacing:.5px;margin-bottom:5px">Format</div>
                      <div style="display:flex;gap:5px" id="fmtBtns_{{cc.idx}}">
                        {% for f,lbl,icon in [('text','Text','Aa'),('number','Number','#'),('alphanumeric','Alpha-Num','A1'),('dropdown','Dropdown','▾')] %}
                        <button type="button"
                          class="type-btn {{'selected-fillable' if (cc.get('format','text')==f) else ''}}"
                          id="fb_{{cc.idx}}_{{f}}"
                          style="padding:4px 9px;font-size:11px"
                          onclick="setFmt({{cc.idx}},'{{f}}')">
                          <span style="font-family:monospace;font-weight:800;font-size:10px">{{icon}}</span>&nbsp;{{lbl}}
                        </button>
                        {% endfor %}
                        <input type="hidden" name="fmt_{{cc.idx}}" value="{{cc.get('format','text')}}" id="fmtHid_{{cc.idx}}"/>
                      </div>
                    </div>

                    <!-- Mandatory -->
                    <div>
                      <div style="font-size:9.5px;font-weight:700;color:var(--amber);text-transform:uppercase;
                                  letter-spacing:.5px;margin-bottom:5px">Required</div>
                      <div style="display:flex;gap:5px">
                        <button type="button"
                          class="type-btn {{'selected-fillable' if cc.get('required') else ''}}"
                          id="reqY_{{cc.idx}}"
                          onclick="setReq({{cc.idx}},true)"
                          style="padding:4px 10px;font-size:11px">
                          OK Yes
                        </button>
                        <button type="button"
                          class="type-btn {{'selected-fillable' if not cc.get('required') else ''}}"
                          id="reqN_{{cc.idx}}"
                          onclick="setReq({{cc.idx}},false)"
                          style="padding:4px 10px;font-size:11px">
                          ⬜ Optional
                        </button>
                        <input type="hidden" name="req_{{cc.idx}}" value="{{'1' if cc.get('required') else '0'}}" id="reqHid_{{cc.idx}}"/>
                      </div>
                    </div>

                    <!-- Dropdown options (only shown when format=dropdown) -->
                    <div id="ddOpts_{{cc.idx}}"
                         style="display:{{'block' if cc.get('format')=='dropdown' else 'none'}};flex:1;min-width:200px">
                      <div style="font-size:9.5px;font-weight:700;color:var(--amber);text-transform:uppercase;
                                  letter-spacing:.5px;margin-bottom:5px">Dropdown Options
                        <span style="font-weight:400;text-transform:none;font-size:9px">
                          (comma-separated, e.g. Yes,No,NA)
                        </span>
                      </div>
                      <input type="text" name="opts_{{cc.idx}}" id="optsInp_{{cc.idx}}"
                             class="fi" placeholder="Yes, No, NA"
                             value="{{ cc.get('options',[])|join(', ') }}"
                             style="font-size:12px;padding:6px 10px;max-width:300px"/>
                    </div>

                  </div>
                </div>
                {% endif %}

              </div><!-- /col-card -->
              {% endfor %}

              <div style="margin-top:18px;display:flex;gap:10px;justify-content:flex-end">
                <a href="/upload" class="btn btn-outline">Cancel</a>
                <button type="submit" class="btn btn-success" id="saveBtn">
                  OK Save & Process →
                </button>
              </div>
            </form>
          </div>
        </div>
      </div>

      <!-- Sticky summary sidebar -->
      <div>
        <div class="card" style="position:sticky;top:70px">
          <div class="card-head">
            <div class="card-icon ci-c"></div>
            <div><h2>Summary</h2><p>Live counts</p></div>
          </div>
          <div class="card-body">
            <div style="margin-bottom:10px">
              <div style="font-size:9.5px;color:var(--text3);font-weight:700;text-transform:uppercase;letter-spacing:.5px">🔒 Locked</div>
              <div style="font-size:24px;font-weight:800;color:#1e293b" id="cntLocked">0</div>
            </div>
            <div style="margin-bottom:10px">
              <div style="font-size:9.5px;color:var(--amber);font-weight:700;text-transform:uppercase;letter-spacing:.5px">✏️ Fillable</div>
              <div style="font-size:24px;font-weight:800;color:var(--amber)" id="cntFillable">0</div>
              <div id="fmtSummary" style="font-size:10px;color:var(--text3);margin-top:3px;line-height:1.8"></div>
            </div>
            <div style="margin-bottom:14px">
              <div style="font-size:9.5px;color:var(--purple);font-weight:700;text-transform:uppercase;letter-spacing:.5px">📎 PDF</div>
              <div style="font-size:24px;font-weight:800;color:var(--purple)" id="cntPdf">0</div>
            </div>
            <div style="padding-top:12px;border-top:1px solid var(--border);font-size:11px;color:var(--text3)">
              <div style="margin-bottom:3px"><strong>File:</strong> {{pending.filename}}</div>
              <div style="margin-bottom:3px"><strong>Datasheet:</strong> {{pending.week}}</div>
              <div style="margin-bottom:3px"><strong>Rows:</strong> {{pending.data|length}}</div>
              <div><strong>Columns:</strong> {{pending.col_cfg|length}}</div>
            </div>
          </div>
        </div>

        <!-- Format legend card -->
        <div class="card" style="margin-top:0">
          <div class="card-head">
            <div class="card-icon ci-a">📖</div>
            <div><h2>Format Guide</h2></div>
          </div>
          <div class="card-body" style="padding:12px 14px">
            <div style="font-size:11px;color:var(--text3);line-height:2">
              <div><span style="font-family:monospace;font-weight:800;color:#1e293b">Aa</span> <strong>Text</strong> — any characters</div>
              <div><span style="font-family:monospace;font-weight:800;color:#1e293b">#</span> <strong>Number</strong> — digits only (commas ok)</div>
              <div><span style="font-family:monospace;font-weight:800;color:#1e293b">A1</span> <strong>Alpha-Num</strong> — letters &amp; numbers</div>
              <div><span style="font-family:monospace;font-weight:800;color:#1e293b">▾</span> <strong>Dropdown</strong> — pick from a list</div>
              <div style="margin-top:6px;padding-top:6px;border-top:1px solid var(--border)">
                <strong>OK Required</strong> — branch must fill before submit<br>
                <strong>⬜ Optional</strong> — can leave blank
              </div>
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>
</div></div>
<script>
function setType(idx, type){
  ['locked','fillable','pdf'].forEach(t=>{
    const btn=document.getElementById('tb_'+idx+'_'+t);
    if(btn) btn.className='type-btn'+(t===type?' selected-'+t:'');
  });
  document.getElementById('hid_'+idx).value=type;
  // Show/hide options panel
  const opts=document.getElementById('opts_'+idx);
  if(opts) opts.style.display = type==='fillable'?'block':'none';
  // Update card border colour
  const card=document.getElementById('card_'+idx);
  if(card){
    card.style.borderColor = type==='fillable'?'#2563eb':type==='pdf'?'#7c3aed':'#e2e8f0';
  }
  updateCounts();
}

function setFmt(idx, fmt){
  ['text','number','alphanumeric','dropdown'].forEach(f=>{
    const btn=document.getElementById('fb_'+idx+'_'+f);
    if(btn) btn.className='type-btn'+(f===fmt?' selected-fillable':'');
  });
  document.getElementById('fmtHid_'+idx).value=fmt;
  // Show/hide dropdown options input
  const dd=document.getElementById('ddOpts_'+idx);
  if(dd) dd.style.display=fmt==='dropdown'?'block':'none';
  updateCounts();
}

function setReq(idx, required){
  document.getElementById('reqY_'+idx).className='type-btn'+(required?' selected-fillable':'');
  document.getElementById('reqN_'+idx).className='type-btn'+(required?'':' selected-fillable');
  document.getElementById('reqHid_'+idx).value=required?'1':'0';
}

function updateCounts(){
  let l=0,f=0,p=0;
  const fmtCount={text:0,number:0,alphanumeric:0,dropdown:0};
  let req=0;
  document.querySelectorAll('[id^="hid_"]').forEach(el=>{
    if(el.value==='locked')l++;
    else if(el.value==='fillable'){
      f++;
      const idx=el.id.replace('hid_','');
      const fmtEl=document.getElementById('fmtHid_'+idx);
      if(fmtEl) fmtCount[fmtEl.value]=(fmtCount[fmtEl.value]||0)+1;
      const reqEl=document.getElementById('reqHid_'+idx);
      if(reqEl&&reqEl.value==='1') req++;
    }
    else if(el.value==='pdf')p++;
  });
  document.getElementById('cntLocked').textContent=l;
  document.getElementById('cntFillable').textContent=f;
  document.getElementById('cntPdf').textContent=p;
  // Show format breakdown
  const lines=[];
  if(fmtCount.text)        lines.push(`Aa Text: ${fmtCount.text}`);
  if(fmtCount.number)      lines.push(`# Number: ${fmtCount.number}`);
  if(fmtCount.alphanumeric)lines.push(`A1 Alpha-Num: ${fmtCount.alphanumeric}`);
  if(fmtCount.dropdown)    lines.push(`▾ Dropdown: ${fmtCount.dropdown}`);
  if(req)                  lines.push(`OK Required: ${req}`);
  document.getElementById('fmtSummary').innerHTML=lines.join('<br>');
}

document.addEventListener('DOMContentLoaded', updateCounts);
document.getElementById('mapForm').addEventListener('submit',()=>{
  document.getElementById('saveBtn').textContent=' Processing...';
  document.getElementById('saveBtn').disabled=true;
});
</script>
"""

# ══════════════════════════════════════════════════════════════════════════════
# TEMPLATE: DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
T_DASHBOARD = CSS + """<div class="layout">""" + NAV + """
<div class="main">
  <div class="topbar">
    <div><div class="topbar-title">Dashboard</div>
         <div class="topbar-sub">ZO → RO → Branch drill-down</div></div>
    <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap">
      <form method="GET" style="display:flex;gap:7px;align-items:center">
        <label style="font-size:11px;font-weight:600;color:var(--text3)">Datasheet:</label>
        <select name="week" class="fi" style="padding:6px 10px;font-size:12px;width:auto"
                onchange="this.form.submit()">
          <option value="">Select datasheet...</option>
          {% for w in weeks %}<option value="{{w}}" {{'selected' if w==week}}>{{w}}</option>{% endfor %}
        </select>
      </form>
      <a href="/upload" class="btn btn-primary btn-sm"> Upload Data</a>
    </div>
  </div>
  <div class="page">""" + FLASH + """
    {% if not weeks %}
    <div class="empty"><div class="ei"></div><h3>No data yet</h3>
      <p>Upload your Excel datasheet to get started.</p><br>
      <a href="/upload" class="btn btn-primary"> Upload Now</a></div>
    {% else %}

    <!-- KPI strip -->
    <div class="stats-row">
      <div class="stat"><div class="stat-lbl">Zones (ZO)</div>
        <div class="stat-val sv-p">{{kpi.zo}}</div></div>
      <div class="stat"><div class="stat-lbl">Regions (RO)</div>
        <div class="stat-val sv-c">{{kpi.ro}}</div></div>
      <div class="stat"><div class="stat-lbl">Branches</div>
        <div class="stat-val sv-c">{{kpi.branches}}</div></div>
      <div class="stat"><div class="stat-lbl">Total Accounts</div>
        <div class="stat-val sv-c">{{kpi.accounts}}</div></div>
      <div class="stat"><div class="stat-lbl">OK Submitted</div>
        <div class="stat-val sv-e">{{kpi.submitted}}</div>
        <div class="stat-sub">{{kpi.sub_pct}}% done</div></div>
      <div class="stat"><div class="stat-lbl"> Pending</div>
        <div class="stat-val sv-a">{{kpi.pending}}</div></div>
    </div>

    <!-- ZO → RO → Branch accordion -->
    <div class="card">
      <div class="card-head">
        <div class="card-icon ci-c">🗂️</div>
        <div><h2>Zone → Region → Branch</h2><p>Click any level to expand · click branch to see accounts</p></div>
        <div class="card-head-r">
          <button class="btn btn-xs btn-outline" onclick="toggleAll(true)">Expand All</button>
          <button class="btn btn-xs btn-outline" onclick="toggleAll(false)">Collapse All</button>
          <a href="/reminders?week={{week}}" class="btn btn-xs btn-amber"> Reminders</a>
          <a href="/export?week={{week}}" class="btn btn-xs btn-success">Download Export</a>
        </div>
      </div>
      <div class="card-body" style="padding:14px">
        {% if tree %}
        {% for zo, ros in tree.items() %}
        {% set zo_total = namespace(v=0) %}{% set zo_sub = namespace(v=0) %}{% set zo_br = namespace(v=0) %}
        {% for ro2,brs2 in ros.items() %}{% for bc2,b2 in brs2.items() %}
          {% set zo_total.v = zo_total.v + b2.total %}{% set zo_sub.v = zo_sub.v + b2.submitted %}{% set zo_br.v = zo_br.v + 1 %}
        {% endfor %}{% endfor %}
        {% set zo_pct = ((zo_sub.v/zo_total.v*100)|round|int) if zo_total.v else 0 %}
        <div class="zo-block">
          <div class="zo-head" onclick="toggle(this)">
            <span style="font-size:16px">🏢</span>
            <span style="flex:1">{{zo}}</span>
            <span class="badge {{'b-e' if zo_pct==100 else 'b-a' if zo_pct>50 else 'b-r'}}"
                  style="margin-right:8px">{{zo_pct}}%</span>
            <span style="font-size:11px;color:var(--text3);margin-right:8px">
              {{zo_br.v}} branches
            </span>
            <span style="color:var(--text3);font-size:12px">▼</span>
          </div>
          <div class="zo-body">
            {% for ro, branches in ros.items() %}
            {% set ro_total = branches.values()|sum(attribute='total') %}
            {% set ro_sub   = branches.values()|sum(attribute='submitted') %}
            {% set ro_pct   = ((ro_sub/ro_total*100)|round|int) if ro_total else 0 %}
            <div class="ro-block">
              <div class="ro-head" onclick="toggle(this)">
                <span>📍</span>
                <span style="flex:1">{{ro}}</span>
                <span class="badge {{'b-e' if ro_pct==100 else 'b-a' if ro_pct>50 else 'b-r'}}"
                      style="margin-right:6px">{{ro_pct}}%</span>
                <span style="font-size:11px;color:var(--text3);margin-right:6px">{{branches|length}} branches</span>
                <span style="color:var(--text3);font-size:11px">▼</span>
              </div>
              <div class="ro-body">
                {% for bc, b in branches.items() %}
                <div class="br-row" onclick="showBranch('{{bc}}','{{week}}','{{b.name}}')">
                  <span class="bc-tag">{{bc}}</span>
                  <span style="flex:1;font-size:12.5px;font-weight:500">{{b.name}}</span>
                  <div class="prog" style="width:130px">
                    <div class="prog-bar">
                      <div class="prog-fill" style="width:{{b.pct}}%;
                        background:{{'var(--emerald)' if b.status=='Submitted' else 'var(--amber)' if b.status=='Partial' else 'var(--rose)'}}"></div>
                    </div>
                    <span class="prog-pct">{{b.pct}}%</span>
                  </div>
                  <span class="badge {{'b-e' if b.status=='Submitted' else 'b-a' if b.status=='Partial' else 'b-r'}}">
                    {{b.status}}</span>
                  <span style="color:var(--text3);font-size:12px;margin-left:4px">›</span>
                </div>
                {% endfor %}
              </div>
            </div>
            {% endfor %}
          </div>
        </div>
        {% endfor %}
        {% else %}
        <div class="empty" style="padding:32px">
          <div class="ei"></div><p>No data for selected datasheet.</p></div>
        {% endif %}
      </div>
    </div>
    {% endif %}
  </div>
</div></div>

<!-- Branch accounts modal -->
<div class="modal-bg" id="brModal">
  <div class="modal-box">
    <div class="modal-head">
      <div>
        <div style="font-size:14px;font-weight:700" id="mTitle">Branch Accounts</div>
        <div style="font-size:11px;color:var(--text3)" id="mSub"></div>
      </div>
      <button class="modal-close" onclick="document.getElementById('brModal').classList.remove('open')">✕</button>
    </div>
    <div id="mBody" style="padding:16px">Loading...</div>
  </div>
</div>

<script>
function toggle(head){
  head.classList.toggle('open');
  const body=head.nextElementSibling;
  body.style.display=body.style.display==='block'?'none':'block';
  head.querySelector('span:last-child').textContent=body.style.display==='block'?'▲':'▼';
}
function toggleAll(open){
  document.querySelectorAll('.zo-head,.ro-head').forEach(h=>{
    const body=h.nextElementSibling;
    body.style.display=open?'block':'none';
    h.classList.toggle('open',open);
    const arr=h.querySelector('span:last-child');
    if(arr)arr.textContent=open?'▲':'▼';
  });
}
function showBranch(bc, week, name){
  const modal = document.getElementById('brModal');
  document.getElementById('mTitle').innerHTML =
    '<span style="font-weight:700;color:var(--text)">'+name+'</span>'+
    ' <span style="font-size:11px;color:var(--text3);font-weight:500">('+bc+')</span>';
  document.getElementById('mSub').textContent = 'Loading…';
  document.getElementById('mBody').innerHTML =
    '<div style="text-align:center;padding:48px;color:var(--text3)">'+
    '<div style="font-size:36px;margin-bottom:10px;opacity:.3"></div>'+
    '<div>Fetching accounts…</div></div>';
  modal.classList.add('open');

  fetch('/api/branch_accounts?bc='+encodeURIComponent(bc)+'&week='+encodeURIComponent(week))
    .then(r=>r.json()).then(data=>{
      const accs=data.accounts||[];
      document.getElementById('mSub').textContent=
        accs.length+' account'+(accs.length!==1?'s':'')+' · '+"Datasheet: "+week;
      if(!accs.length){
        document.getElementById('mBody').innerHTML=
          '<div class="empty" style="padding:36px"><span class="ei"></span>'+
          '<h3>No accounts</h3></div>';
        return;
      }
      const ST={
        Submitted:{cls:'b-e',icon:'OK'},Pending:{cls:'b-a',icon:''},
        Draft:{cls:'b-c',icon:'💾'},Overdue:{cls:'b-r',icon:'🔴'}
      };
      const F={
        locked  :{bg:'#f5f7fa',bdr:'#dce1ed',lbl:'#8898b4',val:'#18243d'},
        fillable:{bg:'#fffbeb',bdr:'#fde68a',lbl:'#92400e',val:'#451a03'},
        pdf     :{bg:'#eef2ff',bdr:'#c7d2fe',lbl:'#4338ca',val:'#312e81'}
      };
      let html='<div style="padding:16px;display:flex;flex-direction:column;gap:12px">';
      accs.forEach((acc,i)=>{
        const rd=acc.row_data||{},bd=acc.branch_data||{},cfg=acc.col_config||[];
        const st=ST[acc.status]||ST.Pending;
        const fields=cfg.filter(cc=>!['zo','ro','branch_code','branch_name'].includes(cc.known));
        const done=fields.filter(cc=>(cc.type==='fillable'?bd[cc.name]:rd[cc.name])).length;
        html+=
          '<div style="border:1.5px solid var(--border);border-radius:12px;overflow:hidden;'+
          'box-shadow:0 1px 4px rgba(24,36,61,.05)">'+
          '<div style="background:var(--surface2);padding:10px 16px;border-bottom:1.5px solid var(--border);'+
          'display:flex;align-items:center;gap:10px">'+
            '<div style="width:26px;height:26px;border-radius:7px;background:var(--sky-lt);color:var(--sky2);'+
            'display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:800;flex-shrink:0">'+(i+1)+'</div>'+
            '<span class="badge '+st.cls+'">'+st.icon+' '+acc.status+'</span>'+
            (acc.submitted_at?'<span style="font-size:10.5px;color:var(--text3)">📅 '+acc.submitted_at+'</span>':
              '<span style="font-size:10.5px;color:var(--text3)">Not submitted</span>')+
            '<span style="margin-left:auto;font-size:10.5px;font-weight:600;color:'+
            (done===fields.length&&fields.length>0?'var(--teal2)':'var(--text3)')+'">'+done+'/'+fields.length+' filled</span>'+
          '</div>'+
          '<div style="padding:14px;background:var(--surface);'+
          'display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:8px">';
        fields.forEach(cc=>{
          const raw=cc.type==='fillable'?bd[cc.name]:rd[cc.name];
          const val=raw&&String(raw).trim()?String(raw).trim():null;
          const f=F[cc.type]||F.locked;
          const fmtTag=cc.type==='fillable'&&cc.format&&cc.format!=='text'
            ?'<span style="font-size:8.5px;font-weight:700;padding:1px 5px;border-radius:3px;'+
              'background:'+f.bdr+';color:'+f.lbl+'">'+({number:'#',dropdown:'▾',alphanumeric:'A1'}[cc.format]||'')+'</span>':''
          const reqTag=cc.required?'<span style="color:var(--rose2);font-weight:700">✱</span>':'';
          html+=
            '<div style="background:'+f.bg+';border:1.5px solid '+f.bdr+
            ((!val&&cc.type==='fillable')?';border-style:dashed':'')+
            ';border-radius:9px;padding:10px 12px">'+
              '<div style="display:flex;align-items:center;gap:3px;margin-bottom:5px">'+
                '<span style="font-size:9px;font-weight:700;color:'+f.lbl+
                ';text-transform:uppercase;letter-spacing:.6px;flex:1;line-height:1.3">'+cc.name+'</span>'+
                reqTag+fmtTag+
              '</div>'+
              '<div style="font-size:13px;font-weight:'+(!val?'400':'600')+';color:'+(!val?f.lbl:f.val)+
              ';word-break:break-all;font-style:'+(!val?'italic':'normal')+'">'+
                (cc.type==='pdf'&&val?'📎 '+val.replace(/^[^_]+_[^_]+_[^_]+_/,''):val||'— not filled —')+
              '</div>'+
            '</div>';
        });
        html+='</div></div>';
      });
      html+='</div>';
      document.getElementById('mBody').innerHTML=html;
    }).catch(()=>{
      document.getElementById('mBody').innerHTML=
        '<div class="empty" style="padding:32px"><h3>Error loading</h3><p>Try again.</p></div>';
    });
}
document.getElementById('brModal').addEventListener('click',function(e){
  if(e.target===this) this.classList.remove('open');
});
</script>"""

# ══════════════════════════════════════════════════════════════════════════════
# TEMPLATE: EXPORT
# ══════════════════════════════════════════════════════════════════════════════
T_EXPORT = CSS + """<div class="layout">""" + NAV + """
<div class="main">
  <div class="topbar">
    <div><div class="topbar-title">Export Data</div>
         <div class="topbar-sub">Filter and download Excel reports</div></div>
  </div>
  <div class="page">""" + FLASH + """
    <div class="card">
      <div class="card-head"><div class="card-icon ci-v">Download</div>
        <div><h2>Select Filters</h2><p>Then choose what to download</p></div></div>
      <div class="card-body">
        <div style="display:flex;gap:12px;flex-wrap:wrap;margin-bottom:18px">
          {% for fid,lbl,opts,placeholder in [
            ('fw','Datasheet',weeks,'All Datasheets'),
            ('fzo','Zone (ZO)',zo_list,'All Zones'),
            ('fro','Region (RO)',ro_list,'All Regions'),
            ('fst','Status',['Submitted','Pending','Draft'],'All Status')] %}
          <div>
            <div style="font-size:10px;font-weight:700;color:var(--text3);margin-bottom:4px;
                        text-transform:uppercase;letter-spacing:.5px">{{lbl}}</div>
            <select id="{{fid}}" class="fi" style="padding:7px 10px;font-size:12px;min-width:150px">
              <option value="">{{placeholder}}</option>
              {% for o in opts %}<option>{{o}}</option>{% endfor %}
            </select>
          </div>
          {% endfor %}
        </div>
        <!-- PDF include toggle -->
        <div style="margin-top:14px;padding:12px 14px;background:var(--purple-lt);
                    border:1.5px solid var(--purple);border-radius:9px;
                    display:flex;align-items:center;gap:12px">
          <label style="display:flex;align-items:center;gap:8px;cursor:pointer;flex:1">
            <input type="checkbox" id="incPdf" style="width:16px;height:16px;cursor:pointer;accent-color:var(--purple)"/>
            <div>
              <div style="font-size:13px;font-weight:700;color:var(--purple)">
                📎 Include Branch-Uploaded PDFs
              </div>
              <div style="font-size:11px;color:var(--text3);margin-top:1px">
                Downloads as a ZIP file containing Excel + all PDF documents organised by ZO / RO / Branch
              </div>
            </div>
          </label>
          <span class="badge b-p" style="flex-shrink:0">Optional</span>
        </div>
        <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:14px;margin-top:14px">
            {% for s,icon,title,desc,cls in [
              ('all','','All Accounts','Every account with all columns','btn-primary'),
              ('sub','✅','Submitted Only','Confirmed accounts only','btn-success'),
              ('pend','','Pending Only','Not yet submitted','btn-amber'),
              ('sum','','ZO/RO Summary','Aggregated count by zone/region','btn-outline')] %}
          <div class="card" style="margin-bottom:0;box-shadow:none">
            <div class="card-body" style="padding:14px;text-align:center">
              <div style="font-size:28px;margin-bottom:6px">{{icon}}</div>
              <div style="font-size:13px;font-weight:700;margin-bottom:3px">{{title}}</div>
              <div style="font-size:11px;color:var(--text3);margin-bottom:12px">{{desc}}</div>
                      <button class="btn {{cls}} btn-sm" style="width:100%;justify-content:center"
                        onclick="dl('{{s}}')">Export</button>
            </div>
          </div>
          {% endfor %}
        </div>
        <div style="background:var(--navy);border-radius:10px;padding:14px 18px;
                    display:flex;align-items:center;justify-content:space-between">
          <div style="color:#fff">
            <div style="font-weight:700;margin-bottom:2px">📦 All 4 Sheets in One File</div>
            <div style="font-size:11px;color:rgba(255,255,255,.45)">All Accounts + Submitted + Pending + Summary</div>
          </div>
          <button class="btn btn-success" onclick="dl('multi')">Download Export All (4 Sheets)</button>
        </div>
      </div>
    </div>
  </div>
</div></div>
<script>
function dl(sheet){
  const p=new URLSearchParams({
    week:document.getElementById('fw').value,
    zo:document.getElementById('fzo').value,
    ro:document.getElementById('fro').value,
    status:document.getElementById('fst').value,
    pdfs:document.getElementById('incPdf').checked?'1':'0',
    sheet});
  window.location='/export/download?'+p.toString();
}
{% if request.args.get('week') %}
document.addEventListener('DOMContentLoaded',()=>{
  document.getElementById('fw').value='{{request.args.get("week","")}}';
});{% endif %}
</script>"""

# ══════════════════════════════════════════════════════════════════════════════
# TEMPLATE: SETTINGS
# ══════════════════════════════════════════════════════════════════════════════
T_SETTINGS = CSS + """<div class="layout">""" + NAV + """
<div class="main">
  <div class="topbar">
    <div><div class="topbar-title">Settings</div>
         <div class="topbar-sub">Master data · Admin accounts · Branch management</div></div>
  </div>
  <div class="page">""" + FLASH + """
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:18px">

      <!-- Master upload -->
      <div>
        <div class="card">
          <div class="card-head"><div class="card-icon ci-v">🗺️</div>
            <div><h2>Branch Master File</h2><p>ZO, RO, Branch Code, Name, Email</p></div></div>
          <div class="card-body">
            <div class="al-i alert" style="margin-bottom:12px;font-size:11px">
              ℹ️ Required: RO + Branch Code + Branch Name · Optional: ZO, Email
            </div>
            <form method="POST" enctype="multipart/form-data">
              <input type="hidden" name="action" value="master">
              <div class="uz" style="padding:20px;margin-bottom:12px">
                <input type="file" name="master_file" accept=".xlsx,.xls" required/>
                <div style="font-size:28px;margin-bottom:6px">🗺️</div>
                <div style="font-weight:600;margin-bottom:2px">Drop master Excel here</div>
                <div class="topbar-sub">or click to browse</div>
              </div>
              <button class="btn btn-primary" style="width:100%;justify-content:center">📤 Upload Master</button>
            </form>
            <div style="margin-top:12px;padding-top:12px;border-top:1px solid var(--border)">
              <div style="font-size:12px;font-weight:600;margin-bottom:8px;color:var(--red)">️ Danger Zone</div>
              <form method="POST" onsubmit="return confirm('Clear ALL branch master data? Type YES to confirm.')">
                <input type="hidden" name="action" value="clear_master">
                <div style="display:flex;gap:8px">
                  <input name="confirm_clear" class="fi" type="text" placeholder='Type YES to confirm' style="font-size:12px;padding:7px 10px"/>
                  <button class="btn btn-danger btn-sm"> Clear Master</button>
                </div>
              </form>
            </div>
          </div>
        </div>

        <!-- Branch list -->
        <div class="card">
          <div class="card-head"><div class="card-icon ci-e"></div>
            <div><h2>Branch Master</h2><p>{{branches|length}} branches loaded</p></div>
            <div class="card-head-r">
              <a href="/settings/download_master" class="btn btn-success btn-sm">
                Download Download Master Excel
              </a>
            </div>
          </div>
          <div style="max-height:360px;overflow-y:auto">
            {% if branches %}
            <table><thead><tr><th>Code</th><th>Name</th><th>RO</th><th>ZO</th><th></th></tr></thead>
            <tbody>{% for b in branches %}
            <tr>
              <td><span class="bc-tag">{{b.branch_code}}</span></td>
              <td style="font-size:12px;max-width:120px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{{b.branch_name or '—'}}</td>
              <td style="font-size:11px;color:var(--text3)">{{b.ro or '—'}}</td>
              <td style="font-size:11px;color:var(--text3)">{{b.zo or '—'}}</td>
              <td>
                <div style="display:flex;gap:4px">
                  <form method="POST" style="display:inline" onsubmit="return confirm('Reset password?')">
                    <input type="hidden" name="action" value="reset_branch">
                    <input type="hidden" name="branch_id" value="{{b.id}}">
                    <button class="btn btn-xs btn-outline" title="Reset Password">🔑</button>
                  </form>
                  <form method="POST" style="display:inline" onsubmit="return confirm('Remove branch?')">
                    <input type="hidden" name="action" value="del_branch">
                    <input type="hidden" name="branch_id" value="{{b.id}}">
                    <button class="btn btn-xs btn-danger"></button>
                  </form>
                </div>
              </td>
            </tr>{% endfor %}</tbody></table>
            {% else %}
            <div class="empty" style="padding:24px"><div class="ei"></div><p>No branches yet.</p></div>
            {% endif %}
          </div>
        </div>
      </div>

      <!-- Admin accounts -->
      <div>
        <div class="card">
          <div class="card-head"><div class="card-icon ci-r">👤</div>
            <div><h2>Admin Accounts</h2><p>{{admins|length}} admins</p></div></div>
          <div class="card-body">
            {% for a in admins %}
            <div style="display:flex;align-items:center;gap:10px;padding:9px 12px;
                        background:var(--bg);border-radius:9px;margin-bottom:7px">
              <span style="font-size:16px">{{'👑' if a.id==1 else '👨‍💼'}}</span>
              <div style="flex:1">
                <div style="font-size:12.5px;font-weight:600">{{a.full_name}}</div>
                <div style="font-size:10px;color:var(--text3)">@{{a.username}} · {{a.created_at[:10]}}</div>
              </div>
              {% if a.id==1 %}<span class="badge b-gr">Default</span>
              {% else %}
              <form method="POST" onsubmit="return confirm('Remove admin?')">
                <input type="hidden" name="action" value="del_admin">
                <input type="hidden" name="admin_id" value="{{a.id}}">
                <button class="btn btn-xs btn-danger"></button>
              </form>{% endif %}
            </div>
            {% endfor %}
            <div style="border:1.5px dashed var(--border);border-radius:9px;padding:14px;margin-top:12px">
              <div style="font-size:12px;font-weight:700;margin-bottom:10px">➕ New Admin</div>
              <form method="POST">
                <input type="hidden" name="action" value="add_admin">
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:8px">
                  <div><label class="fl">USERNAME</label>
                    <input name="username" class="fi mono" type="text" placeholder="username" required style="font-size:12px;padding:7px 10px"/></div>
                  <div><label class="fl">FULL NAME</label>
                    <input name="full_name" class="fi" type="text" placeholder="Full Name" style="font-size:12px;padding:7px 10px"/></div>
                </div>
                <div style="margin-bottom:10px"><label class="fl">PASSWORD (min 6)</label>
                  <input name="password" class="fi" type="password" placeholder="Min 6 chars" required style="font-size:12px;padding:7px 10px"/></div>
                <button class="btn btn-primary" style="width:100%;justify-content:center">➕ Create Admin</button>
              </form>
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>
</div></div>"""

# ══════════════════════════════════════════════════════════════════════════════
# TEMPLATE: REMINDERS
# ══════════════════════════════════════════════════════════════════════════════
T_REMINDERS = CSS + """<div class="layout">""" + NAV + """
<div class="main">
  <div class="topbar">
    <div><div class="topbar-title">Send Reminders</div>
         <div class="topbar-sub">Email pending branches via Outlook</div></div>
    <form method="GET" style="display:flex;gap:7px;align-items:center">
      <select name="week" class="fi" style="padding:6px 10px;font-size:12px;width:auto"
              onchange="this.form.submit()">
        <option value="">Select datasheet...</option>
        {% for w in weeks %}<option value="{{w}}" {{'selected' if w==week}}>{{w}}</option>{% endfor %}
      </select>
    </form>
  </div>
  <div class="page">""" + FLASH + """
    {% if pending %}
    <div class="card" style="margin-bottom:12px">
      <div class="card-body" style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:10px">
        <div>
          <div style="font-size:14px;font-weight:700"> {{pending|length}} branches pending</div>
          <div class="topbar-sub">Click individual email buttons or send bulk</div>
        </div>
        <a href="mailto:?bcc={{pending|selectattr('email')|map(attribute='email')|join(',')}}
                 &subject=Data Submission Reminder – {{week}}
                 &body=Dear Branch Manager,%0A%0AThis is a reminder that your data submission for datasheet {{week}} is pending.%0A%0APlease log in to complete.%0A%0ARegards,%0AData Collection Team"
           class="btn btn-amber"> Bulk Email All Pending</a>
      </div>
    </div>
    <div class="card">
      <div class="tbl-wrap"><table>
        <thead><tr><th>Code</th><th>Branch</th><th>ZO</th><th>RO</th>
               <th>Done</th><th>Total</th><th>Progress</th><th>Email</th></tr></thead>
        <tbody>{% for b in pending %}
        <tr>
          <td><span class="bc-tag">{{b.branch_code}}</span></td>
          <td style="font-weight:500">{{b.branch_name}}</td>
          <td style="font-size:11px;color:var(--text3)">{{b.zo or '—'}}</td>
          <td style="font-size:11px;color:var(--text3)">{{b.ro or '—'}}</td>
          <td style="text-align:center;font-weight:600;color:var(--emerald)">{{b.submitted}}</td>
          <td style="text-align:center">{{b.total}}</td>
          <td><div class="prog">
            <div class="prog-bar"><div class="prog-fill" style="width:{{b.pct}}%;background:var(--amber)"></div></div>
            <span class="prog-pct">{{b.pct}}%</span></div></td>
          <td>{% if b.email %}
            <a href="mailto:{{b.email}}?subject=Data Submission Reminder – {{b.branch_code}} – {{week}}&body=Dear Branch Manager ({{b.branch_name}}),%0A%0AYour data submission for datasheet {{week}} is pending.%0ASubmitted: {{b.submitted}} of {{b.total}}.%0A%0APlease log in to complete.%0A%0ARegards,%0AData Collection Team"
               class="btn btn-xs btn-amber"> Email</a>
          {% else %}<span style="font-size:11px;color:var(--text3)">No email</span>{% endif %}</td>
        </tr>{% endfor %}</tbody>
      </table></div>
    </div>
    {% elif week %}
    <div class="empty"><div class="ei">🎉</div><h3>All submitted!</h3>
      <p>No pending branches for {{week}} datasheet.</p></div>
    {% else %}
    <div class="empty"><div class="ei">📅</div><h3>Select a datasheet above</h3></div>
    {% endif %}
  </div>
</div></div>"""

# ══════════════════════════════════════════════════════════════════════════════
# TEMPLATE: BRANCH LOGIN
# ══════════════════════════════════════════════════════════════════════════════
T_BR_LOGIN = CSS + """
<div class="login-page">
  <div class="login-card">
    <div class="login-hero">
      <div class="login-hero-icon"></div>
      <h1>Branch Login</h1>
      <p>Data Collection Portal</p>
    </div>
    <div class="login-body">""" + FLASH + """
      {% if step=='code' %}
        <div style="font-size:16px;font-weight:700;margin-bottom:4px;color:var(--text)">Welcome!</div>
        <div style="font-size:12px;color:var(--text3);margin-bottom:18px">Enter your branch code to continue.</div>
        <form method="POST">
          <input type="hidden" name="action" value="check_code">
          <div class="fg"><label class="fl">BRANCH CODE</label>
            <input name="branch_code" class="fi" style="font-family:'DM Mono',monospace;font-weight:600;letter-spacing:1px"
                   type="text" placeholder="e.g. MUM001"
                   autofocus oninput="this.value=this.value.toUpperCase()" autocomplete="off" required/>
          </div>
          <button class="btn btn-primary" style="width:100%;justify-content:center;padding:11px">Continue →</button>
        </form>

      {% elif step=='set_pw' %}
        <div style="font-size:16px;font-weight:700;margin-bottom:4px;color:var(--text)">Set Your Password</div>
        <div style="font-size:12px;color:var(--text3);margin-bottom:18px">
          First login for <strong>{{bn}}</strong> ({{bc}}).
        </div>
        <form method="POST">
          <input type="hidden" name="action" value="set_pw">
          <div class="fg"><label class="fl">NEW PASSWORD</label>
            <div class="pw-wrap">
              <input name="password" class="fi" id="p1" type="password"
                     placeholder="Min 6 characters" required oninput="chkStr(this.value)"/>
            </div>
            <div class="strength-bar">
              <div class="s-seg" id="s1"></div><div class="s-seg" id="s2"></div>
              <div class="s-seg" id="s3"></div><div class="s-seg" id="s4"></div>
            </div>
          </div>
          <div class="fg"><label class="fl">CONFIRM PASSWORD</label>
            <input name="confirm" class="fi" type="password" placeholder="Repeat password" required/>
          </div>
          <button class="btn btn-success" style="width:100%;justify-content:center;padding:11px">OK Set Password & Login</button>
        </form>
        <button class="btn btn-outline" style="width:100%;justify-content:center;margin-top:8px"
                onclick="history.back()">← Different Branch</button>

      {% elif step=='pw' %}
        <div style="font-size:16px;font-weight:700;margin-bottom:4px;color:var(--text)">Welcome Back!</div>
        <div style="font-size:12px;color:var(--text3);margin-bottom:18px"><strong>{{bn}}</strong> · {{bc}}</div>
        <form method="POST">
          <input type="hidden" name="action" value="pw">
          <div class="fg">
            <label class="fl" style="display:flex;justify-content:space-between">
              PASSWORD
              <button type="button" class="link-btn"
                      onclick="document.querySelector('[name=action]').value='forgot';this.form.submit()">
                Forgot?</button>
            </label>
            <div class="pw-wrap">
              <input name="password" class="fi" id="p3" type="password" placeholder="Your password" autofocus required/>
              <button type="button" class="pw-toggle"
                      onclick="const i=document.getElementById('p3');i.type=i.type==='password'?'text':'password'">👁️</button>
            </div>
          </div>
          <button class="btn btn-primary" style="width:100%;justify-content:center;padding:11px">Login →</button>
        </form>
        <button class="btn btn-outline" style="width:100%;justify-content:center;margin-top:8px"
                onclick="history.back()">← Different Branch</button>

      {% elif step=='forgot' %}
        <div style="font-size:16px;font-weight:700;margin-bottom:18px;color:var(--text)">Reset Password</div>
        <form method="POST">
          <input type="hidden" name="action" value="verify_forgot">
          <div class="fg"><label class="fl">SECURITY QUESTION</label>
            <div style="padding:9px 12px;background:var(--surface2);border:1.5px solid var(--border);
                        border-radius:8px;font-size:13px;color:var(--text2)">
              What is the name of this portal?</div>
          </div>
          <div class="fg"><label class="fl">YOUR ANSWER</label>
            <input name="answer" class="fi" style="font-family:'DM Mono',monospace;font-weight:600"
                   type="text" autofocus oninput="this.value=this.value.toUpperCase()" placeholder="Type answer…" required/>
          </div>
          <button class="btn btn-primary" style="width:100%;justify-content:center">Verify →</button>
        </form>

      {% elif step=='reset_pw' %}
        <div style="font-size:16px;font-weight:700;margin-bottom:4px;color:var(--text)">New Password</div>
        <div style="font-size:12px;color:var(--teal2);margin-bottom:18px">OK Identity verified</div>
        <form method="POST">
          <input type="hidden" name="action" value="reset_pw">
          <div class="fg"><label class="fl">NEW PASSWORD</label>
            <input name="password" class="fi" type="password" placeholder="Min 6 characters" autofocus required/>
          </div>
          <div class="fg"><label class="fl">CONFIRM PASSWORD</label>
            <input name="confirm" class="fi" type="password" placeholder="Repeat" required/>
          </div>
          <button class="btn btn-success" style="width:100%;justify-content:center;padding:11px">OK Reset & Login</button>
        </form>
      {% endif %}

      <div style="margin-top:18px;text-align:center;padding-top:14px;border-top:1px solid var(--border)">
        <a href="/admin/login" style="font-size:11px;color:var(--text3)">Admin Login →</a>
      </div>
    </div>
  </div>
</div>
<script>
function chkStr(v){
  let s=0;
  if(v.length>=6)s++;if(v.length>=9)s++;
  if(/[A-Z]/.test(v)&&/[0-9]/.test(v))s++;
  if(/[^A-Za-z0-9]/.test(v))s++;
  ['s1','s2','s3','s4'].forEach((id,i)=>{
    const el=document.getElementById(id);if(!el)return;
    el.className='s-seg'+(i<s?(s<=1?' w':s<=2?' f':' s'):'');
  });
}
</script>
"""

# ══════════════════════════════════════════════════════════════════════════════
# TEMPLATE: BRANCH HOME (sheet-wise tabs)
# ══════════════════════════════════════════════════════════════════════════════
T_BR_HOME = CSS + """
<div style="min-height:100vh;background:var(--bg)">
  <div class="br-hdr">
    <div>
      <div style="font-size:14px;font-weight:700">{{bn}}</div>
      <div style="font-size:10px;color:rgba(255,255,255,.6)">{{bc}}{% if ro %} - {{ro}}{% endif %}</div>
    </div>
    <div style="display:flex;gap:8px;align-items:center">
      {% if week %}
      <a href="/branch/download?week={{week|urlencode}}" class="btn btn-outline btn-sm"
         style="font-size:11px">Download My Sheet</a>{% endif %}
      <a href="/branch/logout" style="font-size:11px;color:rgba(255,255,255,.8);background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.2);padding:4px 10px;border-radius:6px;text-decoration:none">Sign Out</a>
    </div>
  </div>
  {% if week %}
  <div class="br-hdr-prog"><div class="br-hdr-fill" style="width:{{pct}}%"></div></div>
  {% endif %}

  <div class="br-wrap">
    """ + FLASH + """
    {% if not weeks %}
    <div class="empty" style="margin-top:40px">
      <div class="ei"></div><h3>No accounts yet</h3>
      <p>Central office has not uploaded data for your branch yet.</p></div>
    {% else %}

    <!-- Sheet-wise datasheet tabs -->
    <div style="margin-top:16px;margin-bottom:4px;font-size:10px;font-weight:700;color:var(--text3);text-transform:uppercase;letter-spacing:.6px">Select Datasheet</div>
    <div class="ds-tabs">
      {% for w in weeks %}
      {% set ws = week_stats.get(w, {}) %}
      <a href="/branch/home?week={{w|urlencode}}"
         class="ds-tab {{'active' if w==week}}">
        {{w}}
        <span style="font-size:9px;margin-left:4px;opacity:.7">{{ws.get('pct',0)}}%</span>
      </a>
      {% endfor %}
    </div>

    {% if week %}
    <!-- Stats -->
    <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin:14px 0">
      {% for lbl,val,cls in [('Total',stats.total,'sv-b'),('OK Done',stats.submitted,'sv-g'),
                              (' Pending',stats.pending,'sv-a'),('Progress',pct|string+'%','sv-p')] %}
      <div class="stat" style="padding:12px;text-align:center">
        <div class="stat-lbl">{{lbl}}</div>
        <div class="stat-val {{cls}}" style="font-size:20px">{{val}}</div>
      </div>
      {% endfor %}
    </div>

    {% if stats.pending==0 %}
    <div class="alert al-s">🎉 <strong>All accounts submitted!</strong> Thank you.</div>
    {% else %}
    <div class="alert al-w"> {{stats.pending}} account(s) pending. Please fill and submit.</div>
    {% endif %}

    <!-- Filter buttons -->
    <div style="display:flex;gap:5px;margin-bottom:12px;flex-wrap:wrap">
      {% for f,lbl in [('all','All'),('Pending',' Pending'),('Draft','💾 Draft'),('Submitted','OK Done')] %}
      <button onclick="filt('{{f}}')" id="fb_{{f}}"
              class="btn btn-sm {{'btn-primary' if f=='all' else 'btn-outline'}}"
              style="font-size:11px">{{lbl}}</button>
      {% endfor %}
    </div>

    <!-- Account cards -->
    <div id="accList">
      {% for acc in accounts %}
      {% set cls='subm' if acc.status=='Submitted' else ('draft' if acc.status=='Draft' else 'pend') %}
      {% set ncls='sv-g' if acc.status=='Submitted' else ('sv-b' if acc.status=='Draft' else 'sv-a') %}
      {% set bcls='b-e' if acc.status=='Submitted' else ('b-c' if acc.status=='Draft' else 'b-a') %}
      <a href="/branch/account/{{acc.id}}" class="acc-card {{cls}}" data-status="{{acc.status}}">
        <div style="width:32px;height:32px;border-radius:8px;display:flex;align-items:center;
                    justify-content:center;font-size:12px;font-weight:800;flex-shrink:0;
                    background:{{'var(--green-lt)' if acc.status=='Submitted' else 'var(--blue-lt)' if acc.status=='Draft' else 'var(--amber-lt)'}};
                    color:{{'var(--emerald)' if acc.status=='Submitted' else 'var(--blue)' if acc.status=='Draft' else 'var(--amber)'}}">
          {{loop.index}}</div>
        <div style="flex:1;min-width:0">
          {% for name,val in acc.summary %}
          <div style="font-size:{{'11px' if loop.index>1 else '13px'}};
                      font-weight:{{'500' if loop.index>1 else '600'}};
                      color:{{'var(--text3)' if loop.index>1 else '#1e293b'}};
                      overflow:hidden;text-overflow:ellipsis;white-space:nowrap">
            {% if loop.index>1 %}<span style="color:var(--text3)">{{name}}: </span>{% endif %}{{val}}
          </div>
          {% endfor %}
          {% if not acc.summary %}
          <div style="font-size:13px;font-weight:600;color:var(--text3)">Account #{{loop.index}}</div>{% endif %}
        </div>
        <div style="text-align:right;flex-shrink:0">
          <span class="badge {{bcls}}">{{acc.status}}</span>
          {% if acc.submitted_at %}<div style="font-size:9px;color:var(--text3);margin-top:3px">{{acc.submitted_at[:10]}}</div>{% endif %}
        </div>
        <span style="color:var(--text3);font-size:16px;margin-left:4px">›</span>
      </a>
      {% endfor %}
    </div>
    {% endif %}
    {% endif %}
  </div>
</div>
<script>
function filt(f){
  document.querySelectorAll('[id^="fb_"]').forEach(b=>{
    b.className=b.className.replace('btn-primary','btn-outline');});
  const fb=document.getElementById('fb_'+f);
  if(fb)fb.className=fb.className.replace('btn-outline','btn-primary');
  document.querySelectorAll('#accList .acc-card').forEach(c=>{
    c.style.display=(f==='all'||c.dataset.status===f)?'flex':'none';});
}
</script>"""

# ══════════════════════════════════════════════════════════════════════════════
# TEMPLATE: BRANCH ACCOUNT FORM
# ══════════════════════════════════════════════════════════════════════════════
T_BR_ACCOUNT = CSS + """
<div style="min-height:100vh;background:var(--bg)">
  <div class="br-hdr">
    <div style="display:flex;align-items:center;gap:10px">
      <a href="/branch/home?week={{acc.week_label|urlencode}}"
         style="background:rgba(255,255,255,.1);border:1px solid rgba(255,255,255,.2);
                color:rgba(255,255,255,.8);padding:4px 10px;border-radius:6px;
                font-size:11px;font-weight:600;text-decoration:none">← Back</a>
      <div>
        <div style="font-size:13px;font-weight:700">{{bn}} · Account {{idx}}/{{total}}</div>
        <div style="font-size:10px;color:rgba(255,255,255,.4)">{{pct}}% of week complete · {{acc.week_label}}</div>
      </div>
    </div>
    <a href="/branch/logout" style="font-size:10px;color:rgba(255,255,255,.4);
       background:rgba(255,255,255,.06);border:1px solid rgba(255,255,255,.12);
       padding:4px 10px;border-radius:6px;text-decoration:none">Sign Out</a>
  </div>
  <div class="br-hdr-prog"><div class="br-hdr-fill" style="width:{{pct}}%"></div></div>

  <div class="br-wrap" style="padding-top:14px">
    """ + FLASH + """

    <!-- Navigation strip -->
    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:12px">
      {% if prev_id %}
      <a href="/branch/account/{{prev_id}}" class="btn btn-outline btn-sm">← Prev</a>
      {% else %}<div></div>{% endif %}
      <div style="display:flex;gap:4px">
        {% for i in range(1,total+1) %}
        <div style="width:7px;height:7px;border-radius:50%;
          background:{{'var(--emerald)' if i<idx else 'var(--blue)' if i==idx else 'var(--border)'}}"></div>
        {% endfor %}
      </div>
      {% if next_id %}
      <a href="/branch/account/{{next_id}}" class="btn btn-outline btn-sm">Next →</a>
      {% else %}<div></div>{% endif %}
    </div>

    <form method="POST" enctype="multipart/form-data">
      <div class="card">
        <div class="card-head">
          <div style="flex:1">
            <div style="font-size:14px;font-weight:700;margin-bottom:3px">
              {% if acc.status=='Submitted' %}OK Submitted{% else %}📝 Fill & Submit{% endif %}
            </div>
            <div class="topbar-sub">
              Week: {{acc.week_label}} ·
              {% set fc=cfg|selectattr('type','equalto','fillable')|list %}
              {{fc|length}} field(s) to fill
            </div>
          </div>
          {% if acc.status=='Submitted' %}
          <span class="badge b-e" style="font-size:11px">Submitted {{acc.submitted_at or ''}}</span>
          {% endif %}
        </div>
        <div class="card-body">
          <!-- LEGEND -->
          <div style="display:flex;gap:8px;margin-bottom:16px;flex-wrap:wrap;
                      padding:8px 10px;background:var(--bg);border-radius:8px">
            <span class="tag-locked">🔒 Locked (prefilled)</span>
            <span class="tag-fillable">✏️ Fill this</span>
            <span class="tag-pdf">📎 Upload PDF</span>
          </div>

          <!-- FIELDS — exact Excel columns -->
          {% for cc in cfg %}
          {% if cc.known not in ['zo','ro','branch_code','branch_name'] %}
          <div class="field-row">
            <div class="field-label">
              {{cc.name}}
              {% if cc.type=='locked' %}
                <br><span class="tag-locked">🔒</span>
              {% elif cc.type=='fillable' %}
                <br><span class="tag-fillable">✏️</span>
                {% if cc.get('required') %}<br><span style="font-size:9px;color:var(--red);font-weight:700">✱ REQUIRED</span>{% endif %}
                {% set fmt=cc.get('format','text') %}
                {% if fmt=='number' %}<br><span style="font-size:9px;color:var(--text3);font-family:monospace">#&nbsp;Number</span>
                {% elif fmt=='alphanumeric' %}<br><span style="font-size:9px;color:var(--text3);font-family:monospace">A1&nbsp;Alpha-Num</span>
                {% elif fmt=='dropdown' %}<br><span style="font-size:9px;color:var(--text3);font-family:monospace">▾&nbsp;Dropdown</span>
                {% endif %}
              {% else %}<br><span class="tag-pdf">📎</span>{% endif %}
            </div>
            {% if cc.type=='locked' %}
              {% set val=rd.get(cc.name,'') %}
              <div class="field-locked {{'empty' if not val}}">
                {{val if val else '— not provided —'}}
              </div>
            {% elif cc.type=='fillable' %}
              {% set cur=bd.get(cc.name,'') or rd.get(cc.name,'') %}
              {% set fmt=cc.get('format','text') %}
              {% set req=cc.get('required',False) %}
              {% set opts=cc.get('options',[]) %}
              <div class="field-fillable">
                {% if acc.status=='Submitted' %}
                  <div class="field-locked">{{cur or '—'}}</div>
                {% elif fmt=='dropdown' and opts %}
                  <select name="field_{{cc.idx}}" class="fi"
                          style="border:1.5px solid var(--amber);background:#fffbeb;font-size:13px;padding:8px 12px">
                    <option value="">— Select {{cc.name}} —</option>
                    {% for opt in opts %}
                    <option value="{{opt}}" {{'selected' if cur==opt}}>{{opt}}</option>
                    {% endfor %}
                  </select>
                {% elif fmt=='number' %}
                  <input type="text" name="field_{{cc.idx}}"
                         value="{{cur}}"
                         placeholder="Numbers only…"
                         inputmode="numeric"
                         oninput="this.value=this.value.replace(/[^0-9.,\-]/g,'')"
                         style="border-color:{{'var(--rose)' if req and not cur else 'var(--amber)'}}"/>
                {% elif fmt=='alphanumeric' %}
                  <input type="text" name="field_{{cc.idx}}"
                         value="{{cur}}"
                         placeholder="Letters &amp; numbers only…"
                         oninput="this.value=this.value.replace(/[^A-Za-z0-9\s\-\/]/g,'')"
                         style="border-color:{{'var(--rose)' if req and not cur else 'var(--amber)'}}"/>
                {% else %}
                  <input type="text" name="field_{{cc.idx}}"
                         value="{{cur}}"
                         placeholder="Enter {{cc.name}}…"
                         style="border-color:{{'var(--rose)' if req and not cur else 'var(--amber)'}}"/>
                {% endif %}
                {% if req and not cur and acc.status!='Submitted' %}
                <div style="font-size:10px;color:var(--red);margin-top:3px;font-weight:600">
                  ✱ Required
                </div>
                {% endif %}
                {% if fmt!='text' and fmt!='dropdown' and acc.status!='Submitted' %}
                <div style="font-size:9.5px;color:var(--text3);margin-top:3px">
                  Format: {{'Numbers only' if fmt=='number' else 'Letters &amp; numbers only'}}
                </div>
                {% endif %}
              </div>
            {% else %}  {# PDF #}
              <div class="field-pdf">
                {% set fname=pf.get(cc.name,'') %}
                {% if fname %}
                  <div class="pdf-done">
                    📎 <strong>{{fname.split('_',3)[-1] if '_' in fname else fname}}</strong>
                    {% if acc.status!='Submitted' %}
                    <span style="font-size:10px;color:var(--text3);margin-left:6px">
                      · Upload again to replace</span>{% endif %}
                  </div>
                {% endif %}
                {% if acc.status!='Submitted' %}
                <input type="file" name="pdf_{{cc.idx}}" accept=".pdf,.jpg,.jpeg,.png"
                       style="margin-top:{{'6px' if fname else '0'}}"/>
                <div style="font-size:10px;color:var(--text3);margin-top:3px">PDF / JPG / PNG · max 20MB</div>
                {% endif %}
              </div>
            {% endif %}
          </div>
          {% endif %}
          {% endfor %}

          <!-- System info (zo/ro/code shown at bottom, compact) -->
          <div style="margin-top:14px;padding:10px 14px;background:var(--bg);border-radius:8px;
                      display:flex;gap:14px;flex-wrap:wrap;font-size:11px;color:var(--text3)">
            {% if acc.zo %}<span>🏢 ZO: <strong>{{acc.zo}}</strong></span>{% endif %}
            {% if acc.ro %}<span>📍 RO: <strong>{{acc.ro}}</strong></span>{% endif %}
            <span> Branch: <strong>{{acc.branch_code}}</strong></span>
          </div>
        </div>
      </div>

      {% if acc.status!='Submitted' %}
      <div style="display:flex;gap:10px;margin-bottom:16px">
        <button type="submit" name="_action" value="draft"
                class="btn btn-outline" style="flex:1;justify-content:center;padding:12px">
          💾 Save Draft
        </button>
        <button type="submit" name="_action" value="submit"
                class="btn btn-success" style="flex:2;justify-content:center;padding:12px;font-size:14px">
          OK Submit Account
        </button>
      </div>
      {% endif %}
    </form>

    {% if acc.status!='Submitted' and next_id %}
    <div style="text-align:center">
      <a href="/branch/account/{{next_id}}" class="btn btn-outline btn-sm">Skip to Next →</a>
    </div>
    {% endif %}
  </div>
</div>"""

# ══════════════════════════════════════════════════════════════════════════════
# RUN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    print('\n' + '═'*50)
    print('  Data Collection Portal v3')
    print('  Admin: http://localhost:5000')
    print('  Branch: http://localhost:5000/branch')
    print('  Login: admin / admin123')
    print('═'*50 + '\n')
    app.run(debug=False, host='0.0.0.0', port=5000)
